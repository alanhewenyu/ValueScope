# Copyright (c) 2025-2026 Alan He. Licensed under AGPL-3.0. See LICENSE.
"""DCF Valuation API endpoints."""

from fastapi import APIRouter, HTTPException, Query, Request, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import json as json_mod
import logging
import numpy as np
import os
import time
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed

logger = logging.getLogger("valuescope.valuation")

from modeling.data import (
    validate_ticker,
    _normalize_ticker,
    fetch_company_profile,
    fetch_forex_data,
    get_company_share_float,
    is_a_share,
    _fill_profile_from_financial_data,
    _calculate_beta_akshare,
)
from backend.data_cache import get_historical_financials, get_company_profile as cached_get_profile
from modeling.dcf import (
    calculate_dcf,
    calculate_wacc,
    calculate_buffett,
    get_risk_free_rate,
    reverse_dcf,
    sensitivity_analysis,
    wacc_sensitivity_analysis,
)
from modeling.constants import (
    HISTORICAL_DATA_PERIODS_ANNUAL,
    TERMINAL_RISK_PREMIUM,
    TERMINAL_RONIC_PREMIUM,
)
from backend.cache import get as cache_get, put as cache_put, make_key
from backend.utils import _build_valuation_params

router = APIRouter()


class DCFParams(BaseModel):
    """Valuation parameters for DCF calculation."""
    ticker: str
    apikey: str = ""
    revenue_growth_1: float
    revenue_growth_2: float
    ebit_margin: float
    convergence: float
    revenue_invested_capital_ratio_1: float
    revenue_invested_capital_ratio_2: float
    revenue_invested_capital_ratio_3: float
    tax_rate: Optional[float] = None     # None = use calculated
    wacc: Optional[float] = None         # None = use calculated
    ronic_match_wacc: bool = False


def _safe_json(obj):
    """Convert numpy/pandas types to JSON-serializable Python types."""
    if isinstance(obj, dict):
        return {k: _safe_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_safe_json(v) for v in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        val = float(obj)
        if np.isnan(val) or np.isinf(val):
            return None
        return val
    if isinstance(obj, np.ndarray):
        return _safe_json(obj.tolist())
    if isinstance(obj, float) and (np.isnan(obj) or np.isinf(obj)):
        return None
    return obj


@router.get("/wacc/{ticker}")
def get_wacc(
    ticker: str,
    apikey: str = Query("", description="FMP API key"),
):
    """Calculate WACC for a given ticker."""
    is_valid, err = validate_ticker(ticker)
    if not is_valid:
        raise HTTPException(status_code=400, detail=err)

    normalized = _normalize_ticker(ticker)

    # Check cache first (30 min TTL)
    ck = make_key("wacc", normalized)
    cached = cache_get(ck)
    if cached is not None:
        return cached

    # Fetch data (both cached)
    financial_data = get_historical_financials(
        normalized, "annual", apikey, HISTORICAL_DATA_PERIODS_ANNUAL
    )
    if financial_data is None:
        raise HTTPException(status_code=404, detail=f"Financial data not found: {ticker}")

    profile = cached_get_profile(normalized, apikey)

    summary_df = financial_data["summary"]
    base_year_data = summary_df.iloc[:, 0].copy()
    base_year_data.name = summary_df.columns[0]
    base_year_data["Average Tax Rate"] = financial_data["average_tax_rate"]

    wacc, total_erp, wacc_details = calculate_wacc(
        base_year_data, profile, apikey, verbose=False
    )

    result = _safe_json({
        "wacc": wacc,
        "total_equity_risk_premium": total_erp,
        "details": wacc_details,
        "risk_free_rate": get_risk_free_rate(profile.get("country", "United States")),
    })
    cache_put(ck, result, ttl=1800)  # 30 min
    return result


@router.post("/dcf")
def run_dcf(params: DCFParams):
    """Run DCF valuation with given parameters.

    Returns: DCF results, sensitivity tables, and verdict.
    """
    is_valid, err = validate_ticker(params.ticker)
    if not is_valid:
        raise HTTPException(status_code=400, detail=err)

    normalized = _normalize_ticker(params.ticker)

    # Fetch all needed data (cached)
    financial_data = get_historical_financials(
        normalized, "annual", params.apikey, HISTORICAL_DATA_PERIODS_ANNUAL
    )
    if financial_data is None:
        raise HTTPException(status_code=404, detail=f"Financial data not found: {params.ticker}")

    profile = cached_get_profile(normalized, params.apikey)

    share_info = get_company_share_float(normalized, params.apikey, company_profile=profile)
    summary_df = financial_data["summary"]

    # Prepare base year data
    base_year_col = summary_df.columns[0]
    base_year_data = summary_df.iloc[:, 0].copy()
    base_year_data.name = base_year_col
    base_year_data["Outstanding Shares"] = share_info.get("outstandingShares", 0) or 0
    base_year_data["Average Tax Rate"] = financial_data["average_tax_rate"]
    base_year_data["Revenue Growth (%)"] = summary_df.iloc[
        summary_df.index.get_loc("Revenue Growth (%)"), 0
    ]
    base_year_data["Total Reinvestment"] = summary_df.iloc[
        summary_df.index.get_loc("Total Reinvestment"), 0
    ]

    # TTM detection
    ttm_quarter = financial_data.get("ttm_latest_quarter", "")
    ttm_end_date = financial_data.get("ttm_end_date", "")
    is_ttm = bool(ttm_quarter and ttm_end_date)
    fy_end_month = financial_data.get("fy_end_month", 12)
    base_year = int(str(base_year_col).replace('FY', ''))
    # forecast_year_1: approximate calendar year for Year 1
    # Rule: ending month ≤ 6 → same year; > 6 → next year
    if is_ttm:
        _ttm_end_month = int(ttm_end_date[5:7])
        _ttm_end_year = int(ttm_end_date[:4])
        forecast_year_1 = _ttm_end_year if _ttm_end_month <= 6 else _ttm_end_year + 1
    else:
        forecast_year_1 = base_year if fy_end_month <= 6 else base_year + 1
    ttm_label = f"{base_year_col}{ttm_quarter} TTM" if is_ttm else ""

    # WACC
    risk_free_rate = get_risk_free_rate(profile.get("country", "United States"))
    wacc_calc, total_erp, wacc_details = calculate_wacc(
        base_year_data, profile, params.apikey, verbose=False
    )

    # Use provided or calculated values
    tax_rate = params.tax_rate if params.tax_rate is not None else financial_data["average_tax_rate"] * 100
    wacc_val = params.wacc if params.wacc is not None else wacc_calc * 100

    # RONIC
    if params.ronic_match_wacc:
        ronic = risk_free_rate + TERMINAL_RISK_PREMIUM
    else:
        ronic = risk_free_rate + TERMINAL_RISK_PREMIUM + TERMINAL_RONIC_PREMIUM

    # Build valuation params
    raw_params = {
        "revenue_growth_1": params.revenue_growth_1,
        "revenue_growth_2": params.revenue_growth_2,
        "ebit_margin": params.ebit_margin,
        "convergence": params.convergence,
        "revenue_invested_capital_ratio_1": params.revenue_invested_capital_ratio_1,
        "revenue_invested_capital_ratio_2": params.revenue_invested_capital_ratio_2,
        "revenue_invested_capital_ratio_3": params.revenue_invested_capital_ratio_3,
        "tax_rate": tax_rate,
        "wacc": wacc_val,
        "ronic": ronic,
    }
    valuation_params = _build_valuation_params(
        raw_params, base_year, risk_free_rate, is_ttm, ttm_quarter, ttm_label,
        forecast_year_1=forecast_year_1, fy_end_month=fy_end_month
    )

    # Run DCF
    results = calculate_dcf(
        base_year_data, valuation_params, financial_data, share_info, profile
    )

    # Forex rate
    forex_rate = None
    reported_currency = results.get("reported_currency", "")
    stock_currency = profile.get("currency", "USD")
    if reported_currency and stock_currency and reported_currency != stock_currency:
        try:
            if params.apikey:
                forex_data = fetch_forex_data(params.apikey)
                forex_key = f"{stock_currency}/{reported_currency}"
                rate = forex_data.get(forex_key)
                if rate and rate != 0:
                    forex_rate = 1.0 / rate
                else:
                    reverse_key = f"{reported_currency}/{stock_currency}"
                    reverse_rate = forex_data.get(reverse_key)
                    if reverse_rate and reverse_rate != 0:
                        forex_rate = reverse_rate

            if forex_rate is None:
                from modeling.yfinance_data import fetch_forex_yfinance
                forex_rate = fetch_forex_yfinance(reported_currency, stock_currency)

            if forex_rate is None:
                from modeling.data import fetch_forex_akshare
                forex_rate = fetch_forex_akshare(reported_currency, stock_currency)
        except Exception as e:
            logger.debug("share float fetch failed: %s", e)

    # Sensitivity
    sens_table = sensitivity_analysis(
        base_year_data, valuation_params, financial_data, share_info, profile
    )
    wacc_results, wacc_base = wacc_sensitivity_analysis(
        base_year_data, valuation_params, financial_data, share_info, profile
    )

    # Reverse DCF: solve for market-implied growth rate
    market_price = profile.get("price", 0)
    reverse_dcf_result = None
    if market_price and market_price > 0:
        try:
            reverse_dcf_result = reverse_dcf(
                base_year_data, valuation_params, financial_data, share_info, profile,
                market_price, forex_rate
            )
        except Exception as e:
            logger.debug("Reverse DCF failed: %s", e)

    # Build price per share (with forex if needed)
    dcf_price = results.get("price_per_share", 0)
    if forex_rate and dcf_price:
        dcf_price_converted = dcf_price * forex_rate
    else:
        dcf_price_converted = dcf_price

    # Verdict
    if market_price and dcf_price_converted:
        diff_pct = (dcf_price_converted - market_price) / market_price
    else:
        diff_pct = 0

    return _safe_json({
        "ticker": normalized,
        "company_name": profile.get("companyName", ""),
        "dcf_price": dcf_price,
        "dcf_price_converted": dcf_price_converted,
        "market_price": market_price,
        "currency": stock_currency,
        "reported_currency": reported_currency,
        "forex_rate": forex_rate,
        "diff_pct": diff_pct,
        "valuation_params": valuation_params,
        "wacc": {
            "value": wacc_calc,
            "details": wacc_details,
        },
        "results": {
            "operating_value": results.get("pv_cf_next_10_years", 0) + results.get("pv_terminal_value", 0),
            "equity_value": results.get("equity_value", 0),
            "price_per_share": results.get("price_per_share", 0),
            "enterprise_value": results.get("enterprise_value", 0),
        },
        "bridge": {
            "pv_cashflows": results.get("pv_cf_next_10_years", 0),
            "pv_terminal_value": results.get("pv_terminal_value", 0),
            "cash": results.get("cash", 0),
            "total_investments": results.get("total_investments", 0),
            "total_debt": results.get("total_debt", 0),
            "minority_interest": results.get("minority_interest", 0),
            "outstanding_shares": results.get("outstanding_shares", 0),
        },
        "sensitivity": {
            "growth_margin": {
                "table": sens_table.values.tolist() if hasattr(sens_table, 'values') else sens_table,
                "growth_rates": list(sens_table.columns) if hasattr(sens_table, 'columns') else [],
                "margins": list(sens_table.index) if hasattr(sens_table, 'index') else [],
            },
            "wacc": {
                "results": wacc_results,
                "base": wacc_base,
            },
        },
        "ttm": {
            "is_ttm": is_ttm,
            "label": ttm_label,
            "quarter": ttm_quarter,
        },
        "forecast_table": _dcf_table_to_json(results.get("dcf_table")),
        "reverse_dcf": reverse_dcf_result,
        "buffett": _buffett_to_json(
            summary_df, profile,
            base_year_data.get("Outstanding Shares", 0),
            forex_rate
        ),
    })


@router.get("/buffett/{ticker}")
def get_buffett_valuation(
    ticker: str,
    apikey: str = Query("", description="FMP API key"),
):
    """Standalone Buffett Owner Earnings valuation (no DCF needed)."""
    is_valid, err = validate_ticker(ticker)
    if not is_valid:
        raise HTTPException(status_code=400, detail=err)

    normalized = _normalize_ticker(ticker)
    ck = make_key("buffett", normalized, apikey[:8] if apikey else "nokey")
    cached = cache_get(ck)
    if cached is not None:
        return cached

    financial_data = get_historical_financials(
        normalized, "annual", apikey, HISTORICAL_DATA_PERIODS_ANNUAL
    )
    if financial_data is None:
        raise HTTPException(status_code=404, detail=f"Financial data not found: {ticker}")

    profile = cached_get_profile(normalized, apikey)
    share_info = get_company_share_float(normalized, apikey, company_profile=profile)
    summary_df = financial_data["summary"]
    outstanding_shares = share_info.get("outstandingShares", 0) or 0

    # Forex rate for currency conversion
    forex_rate = None
    reported_currency = summary_df.loc["Reported Currency"].iloc[0] if "Reported Currency" in summary_df.index else "USD"
    stock_currency = profile.get("currency", "USD")
    if reported_currency and stock_currency and reported_currency != stock_currency:
        try:
            if apikey:
                forex_data = fetch_forex_data(apikey)
                forex_key = f"{stock_currency}/{reported_currency}"
                rate = forex_data.get(forex_key)
                if rate and rate != 0:
                    forex_rate = 1.0 / rate
                else:
                    reverse_key = f"{reported_currency}/{stock_currency}"
                    reverse_rate = forex_data.get(reverse_key)
                    if reverse_rate and reverse_rate != 0:
                        forex_rate = reverse_rate
            if forex_rate is None:
                from modeling.yfinance_data import fetch_forex_yfinance
                forex_rate = fetch_forex_yfinance(reported_currency, stock_currency)
            if forex_rate is None:
                from modeling.data import fetch_forex_akshare
                forex_rate = fetch_forex_akshare(reported_currency, stock_currency)
        except Exception as e:
            logger.debug("forex fetch failed for buffett: %s", e)

    result = _buffett_to_json(summary_df, profile, outstanding_shares, forex_rate)
    result["forex_rate"] = forex_rate
    result["reported_currency"] = reported_currency
    result["stock_currency"] = stock_currency
    result["market_price"] = profile.get("price")

    cache_put(ck, result, ttl=3600)
    return result


def _buffett_to_json(summary_df, profile, outstanding_shares, forex_rate):
    """Run Buffett Owner Earnings valuation and return JSON-safe dict."""
    try:
        result = calculate_buffett(summary_df, profile, outstanding_shares, forex_rate)
        # Remove projection table (too verbose for API)
        result.pop('projection', None)
        return _safe_json(result)
    except Exception as e:
        logger.debug("Buffett valuation failed: %s", e)
        return {"available": False, "reason": str(e)}


def _dcf_table_to_json(dcf_table) -> list[dict] | None:
    """Convert DCF forecast DataFrame to JSON-serializable list of row dicts."""
    if dcf_table is None:
        return None
    try:
        import pandas as pd
        if isinstance(dcf_table, pd.DataFrame):
            rows = []
            for _, row in dcf_table.iterrows():
                r = {}
                for col in dcf_table.columns:
                    v = row[col]
                    if isinstance(v, (np.integer,)):
                        r[col] = int(v)
                    elif isinstance(v, (np.floating, float)):
                        r[col] = None if (np.isnan(v) or np.isinf(v)) else round(float(v), 4)
                    else:
                        r[col] = v
                rows.append(r)
            return rows
    except Exception as e:
        logger.debug("share float fetch failed: %s", e)
    return None


@router.get("/dcf-defaults/{ticker}")
def get_dcf_defaults(
    ticker: str,
    apikey: str = Query("", description="FMP API key"),
):
    """Return suggested DCF parameters based on 5-year historical averages.

    For each parameter, returns: suggested value, 5Y average, min, max, and
    per-year values so the frontend can show context (tooltips, ranges).
    """
    is_valid, err = validate_ticker(ticker)
    if not is_valid:
        raise HTTPException(status_code=400, detail=err)

    normalized = _normalize_ticker(ticker)

    financial_data = get_historical_financials(
        normalized, "annual", apikey, HISTORICAL_DATA_PERIODS_ANNUAL
    )
    if financial_data is None:
        raise HTTPException(status_code=404, detail=f"Financial data not found: {ticker}")

    summary_df = financial_data["summary"]
    # summary_df: index = metric names, columns = years (most recent first)
    # We want the historical years (excluding TTM column if present)

    def _extract_row(metric_name: str) -> dict | None:
        """Extract a metric's historical values + stats."""
        if metric_name not in summary_df.index:
            return None
        row = summary_df.loc[metric_name]
        values = {}
        nums = []
        for col in summary_df.columns:
            v = row[col]
            try:
                fv = float(v)
                if not (np.isnan(fv) or np.isinf(fv)):
                    values[str(col)] = round(fv, 2)
                    nums.append(fv)
            except (ValueError, TypeError):
                pass
        if not nums:
            return None
        return {
            "values": values,
            "avg": round(sum(nums) / len(nums), 2),
            "min": round(min(nums), 2),
            "max": round(max(nums), 2),
            "latest": round(nums[0], 2),  # most recent year first
        }

    rev_growth = _extract_row("Revenue Growth (%)")
    ebit_margin = _extract_row("EBIT Margin (%)")
    incremental_margin = _extract_row("Incremental Margin (%)")
    rev_ic = _extract_row("Revenue / IC")
    tax_rate_row = _extract_row("Tax Rate (%)")

    # Suggested defaults: use 5Y averages with sensible fallbacks
    suggested = {
        "revenue_growth_1": round(rev_growth["latest"] if rev_growth else 10, 1),
        "revenue_growth_2": round(rev_growth["avg"] if rev_growth else 5, 1),
        "ebit_margin": round(ebit_margin["avg"] if ebit_margin else 20, 1),
        "convergence": 5,
        "revenue_invested_capital_ratio_1": round(rev_ic["latest"] if rev_ic else 2.0, 1),
        "revenue_invested_capital_ratio_2": round(rev_ic["avg"] if rev_ic else 1.5, 1),
        "revenue_invested_capital_ratio_3": round(
            rev_ic["avg"] * 0.8 if rev_ic else 1.2, 1
        ),  # assume lower efficiency at scale
    }

    # Determine base year and TTM label
    base_year_col = summary_df.columns[0]
    base_year = int(str(base_year_col).replace('FY', ''))
    fy_end_month = financial_data.get("fy_end_month", 12)
    ttm_quarter = financial_data.get("ttm_latest_quarter", "")
    ttm_end_date = financial_data.get("ttm_end_date", "")
    is_ttm = bool(ttm_quarter and ttm_end_date)
    if is_ttm:
        _em = int(ttm_end_date[5:7])
        _ey = int(ttm_end_date[:4])
        forecast_year_1 = _ey if _em <= 6 else _ey + 1
    else:
        forecast_year_1 = base_year if fy_end_month <= 6 else base_year + 1
    ttm_label = f"{base_year_col}{ttm_quarter} TTM" if is_ttm else ""

    return _safe_json({
        "suggested": suggested,
        "history": {
            "revenue_growth": rev_growth,
            "ebit_margin": ebit_margin,
            "incremental_margin": incremental_margin,
            "revenue_ic": rev_ic,
            "tax_rate": tax_rate_row,
        },
        "average_tax_rate": financial_data.get("average_tax_rate"),
        "base_year": base_year,
        "base_year_label": str(base_year_col),
        "forecast_year_1": forecast_year_1,
        "fy_end_month": fy_end_month,
        "ttm_label": ttm_label,
        "ttm_end_date": ttm_end_date or "",
    })


class AIAnalyzeParams(BaseModel):
    """Parameters for AI-powered DCF analysis."""
    ticker: str
    apikey: str = ""
    serper_key: str = ""
    deepseek_key: str = ""


# Rate limiting for free AI quota (server env var keys)
# Persisted to a JSON file so limits survive server restarts.
_AI_DAILY_LIMIT = int(os.environ.get("VS_AI_DAILY_LIMIT", "3"))
_RATE_LIMIT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", ".rate_limits.json"
)
_ai_usage: dict[str, list[float]] = {}
_rate_lock = threading.Lock()


def _load_rate_limits():
    """Load rate limit data from disk."""
    global _ai_usage
    try:
        with open(_RATE_LIMIT_FILE, "r") as f:
            _ai_usage = json_mod.load(f)
    except (FileNotFoundError, json_mod.JSONDecodeError):
        _ai_usage = {}


def _save_rate_limits():
    """Persist rate limit data to disk."""
    try:
        with open(_RATE_LIMIT_FILE, "w") as f:
            json_mod.dump(_ai_usage, f)
    except OSError:
        pass


# Load on module init
_load_rate_limits()


def _get_client_ip(request: Request) -> str:
    """Extract client IP from X-Forwarded-For header or request.client.host."""
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _check_rate_limit(ip: str) -> bool:
    """Check if IP is within daily rate limit. Returns True if allowed."""
    now = time.time()
    day_ago = now - 86400
    with _rate_lock:
        if ip in _ai_usage:
            _ai_usage[ip] = [t for t in _ai_usage[ip] if t > day_ago]
        else:
            _ai_usage[ip] = []
        return len(_ai_usage[ip]) < _AI_DAILY_LIMIT


def _record_usage(ip: str):
    """Record a usage timestamp for the given IP."""
    with _rate_lock:
        if ip not in _ai_usage:
            _ai_usage[ip] = []
        _ai_usage[ip].append(time.time())
        _save_rate_limits()


@router.get("/ai-quota")
def ai_quota(request: Request):
    """Return remaining free AI analysis quota for this client."""
    ip = _get_client_ip(request)
    now = time.time()
    day_ago = now - 86400
    with _rate_lock:
        used = len([t for t in _ai_usage.get(ip, []) if t > day_ago])
    return {"limit": _AI_DAILY_LIMIT, "used": used, "remaining": max(0, _AI_DAILY_LIMIT - used)}


@router.post("/ai-analyze")
def ai_analyze(params: AIAnalyzeParams, request: Request):
    """Run AI analysis to generate recommended DCF parameters.

    Priority:
    1. CLI engines (Claude/Gemini/Qwen) — preferred, with quick auth verification
    2. Cloud AI (DeepSeek + Serper) as fallback
    3. Return 400 if no AI engine is available
    """
    is_valid, err = validate_ticker(params.ticker)
    if not is_valid:
        raise HTTPException(status_code=400, detail=err)

    normalized = _normalize_ticker(params.ticker)

    # Fetch all needed data
    financial_data = get_historical_financials(
        normalized, "annual", params.apikey, HISTORICAL_DATA_PERIODS_ANNUAL
    )
    if financial_data is None:
        raise HTTPException(status_code=404, detail=f"Financial data not found: {params.ticker}")

    profile = fetch_company_profile(normalized, params.apikey)
    profile = _fill_profile_from_financial_data(profile, financial_data)
    if is_a_share(normalized):
        profile["beta"] = _calculate_beta_akshare(normalized)

    summary_df = financial_data["summary"]
    base_year_col = summary_df.columns[0]
    base_year_data = summary_df.iloc[:, 0].copy()
    base_year_data.name = base_year_col
    base_year_data["Average Tax Rate"] = financial_data["average_tax_rate"]

    # Calculate WACC
    wacc_calc, total_erp, wacc_details = calculate_wacc(
        base_year_data, profile, params.apikey, verbose=False
    )

    # TTM detection
    ttm_quarter = financial_data.get("ttm_latest_quarter", "")
    ttm_end_date = financial_data.get("ttm_end_date", "")
    is_ttm = bool(ttm_quarter and ttm_end_date)
    fy_end_month = financial_data.get("fy_end_month", 12)
    base_year = int(str(base_year_col).replace('FY', ''))
    if is_ttm:
        _em = int(ttm_end_date[5:7])
        _ey = int(ttm_end_date[:4])
        forecast_year_1 = _ey if _em <= 6 else _ey + 1
    else:
        forecast_year_1 = base_year if fy_end_month <= 6 else base_year + 1

    # --- Priority 1: CLI engines (with quick auth verification) ---
    cli_error = None
    try:
        from modeling.ai_analyst import _AI_ENGINE, _detect_ai_engine, _ai_engine_display_name
        import modeling.ai_analyst as _ai_mod
        # _detect_ai_engine() now verifies auth, not just `which`
        if _AI_ENGINE is None:
            _ai_mod._AI_ENGINE = _detect_ai_engine()
        if _ai_mod._AI_ENGINE is not None:
            from modeling.ai_analyst import analyze_company
            ai_result = analyze_company(
                ticker=normalized,
                summary_df=summary_df,
                base_year_data=base_year_data,
                company_profile=profile,
                calculated_wacc=wacc_calc,
                calculated_tax_rate=financial_data["average_tax_rate"],
                base_year=base_year,
                ttm_quarter=ttm_quarter if is_ttm else '',
                ttm_end_date=ttm_end_date if is_ttm else '',
                fy_end_month=fy_end_month,
            )
            if ai_result is not None and ai_result.get("parameters") is not None:
                try:
                    engine = _ai_engine_display_name()
                except Exception:
                    engine = "unknown"
                return _safe_json({
                    "parameters": ai_result["parameters"],
                    "reasoning": ai_result.get("raw_text", ""),
                    "engine": engine,
                })
            cli_error = "CLI returned no parameters"
    except Exception as e:
        cli_error = str(e)

    # --- Priority 2: Cloud AI fallback (DeepSeek + Serper) ---
    serper_key = params.serper_key or os.environ.get("SERPER_API_KEY", "")
    deepseek_key = params.deepseek_key or os.environ.get("DEEPSEEK_API_KEY", "")
    using_server_keys = not params.serper_key and not params.deepseek_key

    if not serper_key or not deepseek_key:
        msg = "No AI engine available."
        if cli_error:
            msg += f" CLI error: {cli_error}."
        msg += " Please configure DeepSeek + Serper API keys in settings."
        raise HTTPException(status_code=400, detail=msg)

    if using_server_keys:
        client_ip = _get_client_ip(request)
        if not _check_rate_limit(client_ip):
            raise HTTPException(
                status_code=429,
                detail=f"Daily AI analysis limit ({_AI_DAILY_LIMIT}) exceeded. "
                       "Configure your own API keys (serper_key + deepseek_key) "
                       "for unlimited usage."
            )

    company_name = profile.get('companyName', normalized)
    country = profile.get('country', 'United States')
    beta = profile.get('beta', 1.0)
    market_cap = profile.get('marketCap', 0)
    financial_table = summary_df.to_string()

    f_year_1 = forecast_year_1
    if is_ttm:
        _ttm_label = f'{base_year_col}{ttm_quarter} TTM'
        ttm_context = f'，数据为 {_ttm_label}（截至 {ttm_end_date} 的最近十二个月）'
        ttm_base_label = f' ({_ttm_label})'
        forecast_year_guidance = (
            f'DCF 预测 Year 1 覆盖从 {ttm_end_date} 起的未来12个月（大致对应 {f_year_1} 日历年）。'
            f'请以 {f_year_1} 年作为 Year 1 的参考年份搜索业绩指引和分析师预期。'
        )
    else:
        ttm_context = ''
        ttm_base_label = ''
        forecast_year_guidance = f'Year 1 对应 {f_year_1} 年。'

    template_args = {
        'ticker': normalized,
        'company_name': company_name,
        'country': country,
        'beta': beta,
        'market_cap': f"{market_cap:,.0f}",
        'calculated_wacc': f"{wacc_calc:.2%}",
        'calculated_tax_rate': f"{financial_data['average_tax_rate']:.2%}",
        'financial_table': financial_table,
        'base_year': base_year,
        'forecast_year_guidance': forecast_year_guidance,
        'search_year': f_year_1,
        'search_year_2': f_year_1 + 1,
        'ttm_context': ttm_context,
        'ttm_base_label': ttm_base_label,
    }

    try:
        from modeling.ai_analyst import cloud_ai_analyze, _parse_structured_parameters
        raw_text = cloud_ai_analyze(
            template_args=template_args,
            serper_key=serper_key,
            deepseek_key=deepseek_key,
        )
        ai_params = _parse_structured_parameters(raw_text)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cloud AI analysis failed: {str(e)}")

    if ai_params is None:
        raise HTTPException(status_code=500, detail="Cloud AI analysis returned no parameters")

    if using_server_keys:
        _record_usage(_get_client_ip(request))

    return _safe_json({
        "parameters": ai_params,
        "reasoning": raw_text,
        "engine": "DeepSeek R1",
    })


@router.post("/ai-analyze-stream")
def ai_analyze_stream(params: AIAnalyzeParams, request: Request):
    """Streaming version of ai-analyze using SSE for progress updates."""

    is_valid, err = validate_ticker(params.ticker)
    if not is_valid:
        raise HTTPException(status_code=400, detail=err)

    normalized = _normalize_ticker(params.ticker)

    def _sse(event: str, data: dict) -> str:
        return f"event: {event}\ndata: {json_mod.dumps(data, ensure_ascii=False)}\n\n"

    def generate():
        yield _sse("progress", {"phase": "fetching", "message": "Fetching data & profile..."})

        # Parallel fetch: financials, profile, and beta (if A-share)
        with ThreadPoolExecutor(max_workers=3) as executor:
            fut_fin = executor.submit(
                get_historical_financials, normalized, "annual", params.apikey, HISTORICAL_DATA_PERIODS_ANNUAL
            )
            fut_prof = executor.submit(fetch_company_profile, normalized, params.apikey)
            fut_beta = None
            if is_a_share(normalized):
                fut_beta = executor.submit(_calculate_beta_akshare, normalized)

        try:
            financial_data = fut_fin.result()
        except Exception as e:
            yield _sse("error", {"message": f"Failed to fetch data: {e}"})
            return
        if financial_data is None:
            yield _sse("error", {"message": f"Financial data not found: {params.ticker}"})
            return

        profile = fut_prof.result()
        profile = _fill_profile_from_financial_data(profile, financial_data)
        if fut_beta is not None:
            profile["beta"] = fut_beta.result()

        summary_df = financial_data["summary"]
        base_year_col = summary_df.columns[0]
        base_year_data = summary_df.iloc[:, 0].copy()
        base_year_data.name = base_year_col
        base_year_data["Average Tax Rate"] = financial_data["average_tax_rate"]

        yield _sse("progress", {"phase": "wacc", "message": "Calculating WACC..."})

        wacc_calc, total_erp, wacc_details = calculate_wacc(
            base_year_data, profile, params.apikey, verbose=False
        )

        # TTM detection
        ttm_quarter = financial_data.get("ttm_latest_quarter", "")
        ttm_end_date = financial_data.get("ttm_end_date", "")
        is_ttm = bool(ttm_quarter and ttm_end_date)
        fy_end_month = financial_data.get("fy_end_month", 12)
        base_year = int(str(base_year_col).replace('FY', ''))
        if is_ttm:
            _em = int(ttm_end_date[5:7])
            _ey = int(ttm_end_date[:4])
            forecast_year_1 = _ey if _em <= 6 else _ey + 1
        else:
            forecast_year_1 = base_year if fy_end_month <= 6 else base_year + 1

        # --- Priority 1: CLI engines (with quick auth verification) ---
        cli_error = None
        try:
            from modeling.ai_analyst import _AI_ENGINE, _detect_ai_engine, _ai_engine_display_name
            import modeling.ai_analyst as _ai_mod
            # _detect_ai_engine() now verifies auth, not just `which`
            if _AI_ENGINE is None:
                _ai_mod._AI_ENGINE = _detect_ai_engine()
            if _ai_mod._AI_ENGINE is not None:
                engine_name = _ai_engine_display_name()
                yield _sse("progress", {"phase": "ai_calling", "message": f"Calling {engine_name}... (may take 1-3 min)"})

                result_q: queue.Queue = queue.Queue()

                def _run_cli():
                    try:
                        from modeling.ai_analyst import analyze_company
                        ai_result = analyze_company(
                            ticker=normalized,
                            summary_df=summary_df,
                            base_year_data=base_year_data,
                            company_profile=profile,
                            calculated_wacc=wacc_calc,
                            calculated_tax_rate=financial_data["average_tax_rate"],
                            base_year=base_year,
                            ttm_quarter=ttm_quarter if is_ttm else '',
                            ttm_end_date=ttm_end_date if is_ttm else '',
                            fy_end_month=fy_end_month,
                        )
                        result_q.put(("ok", ai_result))
                    except Exception as e:
                        result_q.put(("error", str(e)))

                t = threading.Thread(target=_run_cli, daemon=True)
                t.start()

                _cli_phases = [
                    "Searching for earnings guidance & analyst forecasts...",
                    "Analyzing revenue growth trends...",
                    "Evaluating operating margins & profitability...",
                    "Assessing competitive position & industry outlook...",
                    "Reviewing WACC & cost of capital assumptions...",
                    "Synthesizing data into DCF parameters...",
                    "Finalizing parameter recommendations...",
                ]
                elapsed = 0
                while t.is_alive():
                    t.join(timeout=5)
                    elapsed += 5
                    if t.is_alive():
                        mins = elapsed // 60
                        secs = elapsed % 60
                        phase_idx = min(elapsed // 15, len(_cli_phases) - 1)
                        phase_msg = _cli_phases[phase_idx]
                        yield _sse("progress", {
                            "phase": "ai_working",
                            "message": f"{phase_msg} ({mins}m {secs}s)",
                            "elapsed": elapsed,
                        })

                status, value = result_q.get_nowait()
                if status == "ok" and value and value.get("parameters"):
                    try:
                        engine = _ai_engine_display_name()
                    except Exception:
                        engine = "unknown"
                    yield _sse("result", _safe_json({
                        "parameters": value["parameters"],
                        "reasoning": value.get("raw_text", ""),
                        "engine": engine,
                    }))
                    return
                cli_error = value if status == "error" else "CLI returned no parameters"
        except Exception as e:
            cli_error = str(e)

        # --- Priority 2: Cloud AI fallback (DeepSeek + Serper) ---
        serper_key = params.serper_key or os.environ.get("SERPER_API_KEY", "")
        deepseek_key = params.deepseek_key or os.environ.get("DEEPSEEK_API_KEY", "")
        using_server_keys = not params.serper_key and not params.deepseek_key

        if not serper_key or not deepseek_key:
            msg = "No AI engine available."
            if cli_error:
                msg += f" CLI error: {cli_error}."
            msg += " Please configure DeepSeek + Serper API keys in settings."
            yield _sse("error", {"message": msg})
            return

        if using_server_keys:
            client_ip = _get_client_ip(request)
            if not _check_rate_limit(client_ip):
                yield _sse("error", {"message": f"Daily AI limit ({_AI_DAILY_LIMIT}) exceeded."})
                return

        yield _sse("progress", {"phase": "cloud_search", "message": f"CLI failed, switching to DeepSeek R1..."})

        # Build cloud template_args
        company_name = profile.get('companyName', normalized)
        country = profile.get('country', 'United States')
        financial_table = summary_df.to_string()
        f_year_1 = forecast_year_1
        if is_ttm:
            _ttm_label = f'{base_year_col}{ttm_quarter} TTM'
            ttm_context = f'，数据为 {_ttm_label}（截至 {ttm_end_date} 的最近十二个月）'
            ttm_base_label = f' ({_ttm_label})'
            forecast_year_guidance = (
                f'DCF 预测 Year 1 覆盖从 {ttm_end_date} 起的未来12个月（大致对应 {f_year_1} 日历年）。'
                f'请以 {f_year_1} 年作为 Year 1 的参考年份搜索业绩指引和分析师预期。'
            )
        else:
            ttm_context = ''
            ttm_base_label = ''
            forecast_year_guidance = f'Year 1 对应 {f_year_1} 年。'

        template_args = {
            'ticker': normalized, 'company_name': company_name, 'country': country,
            'beta': profile.get('beta', 1.0),
            'market_cap': f"{profile.get('marketCap', 0):,.0f}",
            'calculated_wacc': f"{wacc_calc:.2%}",
            'calculated_tax_rate': f"{financial_data['average_tax_rate']:.2%}",
            'financial_table': financial_table, 'base_year': base_year,
            'forecast_year_guidance': forecast_year_guidance,
            'search_year': f_year_1, 'search_year_2': f_year_1 + 1,
            'ttm_context': ttm_context, 'ttm_base_label': ttm_base_label,
        }

        # Cloud AI with progress callback
        progress_q: queue.Queue = queue.Queue()
        cloud_result_q: queue.Queue = queue.Queue()

        def _progress_cb(phase, msg):
            progress_q.put((phase, msg))

        def _run_cloud():
            try:
                from modeling.ai_analyst import cloud_ai_analyze, _parse_structured_parameters
                raw = cloud_ai_analyze(
                    template_args=template_args,
                    serper_key=serper_key,
                    deepseek_key=deepseek_key,
                    progress_callback=_progress_cb,
                )
                params_out = _parse_structured_parameters(raw)
                cloud_result_q.put(("ok", raw, params_out))
            except Exception as e:
                cloud_result_q.put(("error", str(e), None))

        ct = threading.Thread(target=_run_cloud, daemon=True)
        ct.start()

        while ct.is_alive():
            ct.join(timeout=3)
            while not progress_q.empty():
                phase, msg = progress_q.get_nowait()
                label_map = {
                    'searching': 'Searching',
                    'scraping': 'Reading',
                    'analyzing': 'AI reasoning...',
                    'generating': 'DeepSeek R1 generating...',
                }
                display = label_map.get(phase, phase)
                if msg:
                    display = f"{display}: {msg}" if phase != 'analyzing' else display
                yield _sse("progress", {"phase": f"cloud_{phase}", "message": display})

        status, raw_text, ai_params = cloud_result_q.get_nowait()
        if status != "ok" or ai_params is None:
            yield _sse("error", {"message": f"All AI engines failed. Cloud: {raw_text}"})
            return

        if using_server_keys:
            _record_usage(_get_client_ip(request))

        yield _sse("result", _safe_json({
            "parameters": ai_params,
            "reasoning": raw_text,
            "engine": "DeepSeek R1",
        }))

    return StreamingResponse(generate(), media_type="text/event-stream")


# ── Gap Analysis ─────────────────────────────────────────────────────

class GapAnalyzeParams(BaseModel):
    ticker: str
    apikey: str = ""
    serper_key: str = ""
    deepseek_key: str = ""
    # DCF results
    dcf_price: float
    market_price: float
    valuation_params: dict
    bridge: dict
    forex_rate: Optional[float] = None
    reported_currency: str = ""


@router.post("/gap-analyze")
def gap_analyze(params: GapAnalyzeParams, request: Request):
    """Streaming SSE gap analysis comparing DCF result vs market price."""
    import re
    from datetime import date

    is_valid, err = validate_ticker(params.ticker)
    if not is_valid:
        raise HTTPException(status_code=400, detail=err)

    normalized = _normalize_ticker(params.ticker)

    def _sse(event: str, data: dict) -> str:
        return f"event: {event}\ndata: {json_mod.dumps(data, ensure_ascii=False)}\n\n"

    def generate():
        yield _sse("progress", {"message": "Fetching data & profile..."})

        # Parallel fetch: financials and profile
        with ThreadPoolExecutor(max_workers=2) as executor:
            fut_fin = executor.submit(
                get_historical_financials, normalized, "annual", params.apikey, HISTORICAL_DATA_PERIODS_ANNUAL
            )
            fut_prof = executor.submit(fetch_company_profile, normalized, params.apikey)

        try:
            financial_data = fut_fin.result()
        except Exception as e:
            yield _sse("error", {"message": f"Data fetch failed: {e}"})
            return
        if financial_data is None:
            yield _sse("error", {"message": f"Financial data not found: {params.ticker}"})
            return

        profile = fut_prof.result()
        profile = _fill_profile_from_financial_data(profile, financial_data)

        summary_df = financial_data["summary"]
        base_year_col = summary_df.columns[0]
        base_year = int(str(base_year_col).replace('FY', ''))
        fy_end_month = financial_data.get("fy_end_month", 12)
        ttm_end_date = financial_data.get("ttm_end_date", "")
        is_ttm = bool(financial_data.get("ttm_latest_quarter") and ttm_end_date)

        if is_ttm:
            _em = int(ttm_end_date[5:7])
            _ey = int(ttm_end_date[:4])
            forecast_year_1 = _ey if _em <= 6 else _ey + 1
        else:
            forecast_year_1 = base_year if fy_end_month <= 6 else base_year + 1

        stock_currency = profile.get('currency', 'USD')

        bridge = params.bridge
        results_dict = {
            'price_per_share': params.dcf_price,
            'reported_currency': params.reported_currency or stock_currency,
            'pv_cf_next_10_years': bridge.get('pv_cashflows', 0),
            'pv_terminal_value': bridge.get('pv_terminal_value', 0),
            'enterprise_value': bridge.get('pv_cashflows', 0) + bridge.get('pv_terminal_value', 0)
                               + bridge.get('cash', 0) + bridge.get('total_investments', 0),
            'equity_value': bridge.get('pv_cashflows', 0) + bridge.get('pv_terminal_value', 0)
                           + bridge.get('cash', 0) + bridge.get('total_investments', 0)
                           - bridge.get('total_debt', 0) - bridge.get('minority_interest', 0),
        }

        # --- Priority 1: CLI engines (with quick auth verification) ---
        cli_error = None
        try:
            from modeling.ai_analyst import _AI_ENGINE, _detect_ai_engine, analyze_valuation_gap, _ai_engine_display_name
            import modeling.ai_analyst as _ai_mod
            if _AI_ENGINE is None:
                _ai_mod._AI_ENGINE = _detect_ai_engine()
            if _ai_mod._AI_ENGINE is not None:
                engine_name = _ai_engine_display_name()
                yield _sse("progress", {"message": f"Calling {engine_name} for gap analysis..."})

                result_q: queue.Queue = queue.Queue()

                def _run_cli():
                    try:
                        r = analyze_valuation_gap(
                            ticker=normalized, company_profile=profile, results=results_dict,
                            valuation_params=params.valuation_params, summary_df=summary_df,
                            base_year=base_year, forecast_year_1=forecast_year_1, forex_rate=params.forex_rate,
                        )
                        result_q.put(("ok", r))
                    except Exception as e:
                        result_q.put(("error", str(e)))

                _gap_phases = [
                    "Searching for analyst price targets & consensus...",
                    "Reviewing latest earnings & news impact...",
                    "Analyzing risks, headwinds & growth catalysts...",
                    "Comparing DCF assumptions vs market expectations...",
                    "Evaluating competitive position & moat...",
                    "Synthesizing gap analysis & adjusted price...",
                    "Finalizing analysis...",
                ]
                t = threading.Thread(target=_run_cli, daemon=True)
                t.start()
                elapsed = 0
                while t.is_alive():
                    t.join(timeout=5)
                    elapsed += 5
                    if t.is_alive():
                        mins = elapsed // 60
                        secs = elapsed % 60
                        phase_idx = min(elapsed // 15, len(_gap_phases) - 1)
                        yield _sse("progress", {"message": f"{_gap_phases[phase_idx]} ({mins}m {secs}s)"})

                status, value = result_q.get_nowait()
                if status == "ok" and value is not None:
                    display_text = re.sub(r'\n?\s*ADJUSTED_PRICE:.*$', '', value['analysis_text']).strip()
                    try:
                        engine = _ai_engine_display_name()
                    except Exception:
                        engine = "unknown"
                    yield _sse("result", _safe_json({
                        "analysis_text": display_text,
                        "adjusted_price": value.get('adjusted_price'),
                        "dcf_price": value.get('dcf_price', params.dcf_price),
                        "market_price": value.get('current_price', params.market_price),
                        "gap_pct": value.get('gap_pct', 0),
                        "currency": value.get('currency', stock_currency),
                        "engine": engine,
                    }))
                    return
                cli_error = value if status == "error" else "CLI gap analysis returned None"
        except Exception as e:
            cli_error = str(e)

        # --- Priority 2: Cloud AI fallback (DeepSeek + Serper) ---
        serper_key = params.serper_key or os.environ.get("SERPER_API_KEY", "")
        deepseek_key = params.deepseek_key or os.environ.get("DEEPSEEK_API_KEY", "")
        using_server_keys = not params.serper_key and not params.deepseek_key

        if not serper_key or not deepseek_key:
            msg = "No AI engine available for gap analysis."
            if cli_error:
                msg += f" CLI error: {cli_error}."
            msg += " Please configure DeepSeek + Serper API keys in settings."
            yield _sse("error", {"message": msg})
            return

        if using_server_keys:
            client_ip = _get_client_ip(request)
            if not _check_rate_limit(client_ip):
                yield _sse("error", {"message": f"Daily AI limit ({_AI_DAILY_LIMIT}) exceeded."})
                return

        company_name = profile.get('companyName', normalized)
        country = profile.get('country', 'United States')
        current_price = profile.get('price', params.market_price)
        dcf_price = params.dcf_price
        if params.forex_rate and params.reported_currency and params.reported_currency != stock_currency:
            dcf_price = params.dcf_price * params.forex_rate

        gap_pct = (dcf_price - current_price) / current_price * 100 if current_price else 0
        gap_direction = 'DCF 估值高于市场价，市场可能低估' if gap_pct > 0 else 'DCF 估值低于市场价，市场可能高估'
        today = date.today()

        template_args = {
            'ticker': normalized, 'company_name': company_name, 'country': country,
            'current_price': current_price, 'currency': stock_currency,
            'dcf_price': dcf_price, 'gap_pct': gap_pct, 'gap_direction': gap_direction,
            'revenue_growth_1': params.valuation_params.get('revenue_growth_1', 0),
            'revenue_growth_2': params.valuation_params.get('revenue_growth_2', 0),
            'ebit_margin': params.valuation_params.get('ebit_margin', 0),
            'wacc': params.valuation_params.get('wacc', 0),
            'tax_rate': params.valuation_params.get('tax_rate', 0),
            'pv_cf': results_dict['pv_cf_next_10_years'],
            'pv_terminal': results_dict['pv_terminal_value'],
            'enterprise_value': results_dict['enterprise_value'],
            'equity_value': results_dict['equity_value'],
            'financial_table': summary_df.to_string(),
            'forecast_year': forecast_year_1,
            'current_date': today.strftime('%Y-%m-%d'),
            'current_year': today.year,
        }

        yield _sse("progress", {"message": "CLI failed, switching to DeepSeek R1..."})

        progress_q: queue.Queue = queue.Queue()
        cloud_result_q: queue.Queue = queue.Queue()

        def _progress_cb(phase, msg):
            progress_q.put((phase, msg))

        def _run_cloud():
            try:
                from modeling.ai_analyst import cloud_gap_analyze
                raw = cloud_gap_analyze(
                    template_args=template_args,
                    serper_key=serper_key,
                    deepseek_key=deepseek_key,
                    progress_callback=_progress_cb,
                )
                cloud_result_q.put(("ok", raw))
            except Exception as e:
                cloud_result_q.put(("error", str(e)))

        ct = threading.Thread(target=_run_cloud, daemon=True)
        ct.start()
        while ct.is_alive():
            ct.join(timeout=3)
            while not progress_q.empty():
                phase, msg = progress_q.get_nowait()
                label_map = {'searching': 'Searching', 'scraping': 'Reading', 'analyzing': 'AI reasoning...', 'generating': 'DeepSeek R1 generating...'}
                display = label_map.get(phase, phase)
                if msg and phase not in ('analyzing', 'generating'):
                    display = f"{display}: {msg}"
                yield _sse("progress", {"message": display})

        status, raw_text = cloud_result_q.get_nowait()
        if status != "ok":
            yield _sse("error", {"message": f"All AI engines failed. Cloud: {raw_text}"})
            return

        adjusted_price = None
        price_match = re.search(r'ADJUSTED_PRICE:\s*([\d.,]+)', raw_text)
        if price_match:
            try:
                adjusted_price = float(price_match.group(1).replace(',', ''))
            except ValueError:
                pass

        display_text = re.sub(r'\n?\s*ADJUSTED_PRICE:.*$', '', raw_text).strip()

        if using_server_keys:
            _record_usage(_get_client_ip(request))

        yield _sse("result", _safe_json({
            "analysis_text": display_text,
            "adjusted_price": adjusted_price,
            "dcf_price": dcf_price,
            "market_price": current_price,
            "gap_pct": gap_pct,
            "currency": stock_currency,
            "engine": "DeepSeek R1",
        }))

    return StreamingResponse(generate(), media_type="text/event-stream")


# ── Save Valuation ───────────────────────────────────────────────────

class SaveValuationParams(BaseModel):
    ticker: str
    company_name: str
    mode: str = "manual"
    ai_engine: Optional[str] = None
    valuation_params: dict
    dcf_results: dict
    company_profile: dict
    gap_analysis: Optional[dict] = None
    ai_result: Optional[dict] = None
    sensitivity: Optional[dict] = None
    financial_summary: Optional[dict] = None
    forex_rate: Optional[float] = None


@router.post("/save")
def save_valuation(params: SaveValuationParams, request: Request):
    """Save valuation to SQLite if VS_DB_PATH is set."""
    from datetime import date
    from backend.routers.auth import get_current_user
    user_id = get_current_user(request)

    db_path = os.environ.get("VS_DB_PATH", "")
    if not db_path:
        return {"saved": False, "reason": "no_db"}

    try:
        import json as _json
        import pandas as _pd
        from modeling.db_export import save_to_db

        # Reconstruct results dict from frontend data
        bridge = params.dcf_results.get("bridge", {})
        results = {
            'price_per_share': params.dcf_results.get('dcf_price', 0),
            'reported_currency': params.dcf_results.get('reported_currency', ''),
            'pv_cf_next_10_years': bridge.get('pv_cashflows', 0),
            'pv_terminal_value': bridge.get('pv_terminal_value', 0),
            'enterprise_value': bridge.get('pv_cashflows', 0) + bridge.get('pv_terminal_value', 0)
                               + bridge.get('cash', 0) + bridge.get('total_investments', 0),
            'equity_value': bridge.get('pv_cashflows', 0) + bridge.get('pv_terminal_value', 0)
                           + bridge.get('cash', 0) + bridge.get('total_investments', 0)
                           - bridge.get('total_debt', 0) - bridge.get('minority_interest', 0),
            'cash': bridge.get('cash', 0),
            'total_investments': bridge.get('total_investments', 0),
            'total_debt': bridge.get('total_debt', 0),
            'minority_interest': bridge.get('minority_interest', 0),
            'outstanding_shares': bridge.get('outstanding_shares', 0),
        }

        # Reconstruct forecast table as DataFrame for dcf_table_json
        forecast_rows = params.dcf_results.get("forecast_table")
        if forecast_rows and isinstance(forecast_rows, list):
            try:
                results['dcf_table'] = _pd.DataFrame(forecast_rows)
            except Exception:
                pass

        # Sensitivity: convert frontend format → DataFrames expected by save_to_db
        sens_table = None
        wacc_sens = None
        if params.sensitivity:
            gm = params.sensitivity.get("growth_margin", {})
            if gm:
                table = gm.get("table", [])
                growth_rates = gm.get("growth_rates", [])
                margins = gm.get("margins", [])
                if table and growth_rates and margins:
                    try:
                        sens_table = _pd.DataFrame(
                            table,
                            index=[str(g) for g in growth_rates],
                            columns=[str(m) for m in margins],
                        )
                    except Exception:
                        pass
            ws = params.sensitivity.get("wacc", {})
            if ws:
                wacc_results = ws.get("results", {})
                wacc_base = ws.get("base")
                if wacc_results:
                    wacc_sens = (wacc_results, wacc_base)

        # Reconstruct financial summary as DataFrame for summary_json
        financial_data = None
        if params.financial_summary:
            try:
                cols = params.financial_summary.get("columns", [])
                idx = params.financial_summary.get("index", [])
                data = params.financial_summary.get("data", [])
                if cols and idx and data:
                    summary_df = _pd.DataFrame(data, index=idx, columns=cols)
                    financial_data = {'summary': summary_df}
            except Exception:
                pass

        # Gap analysis: compute adjusted_price_reporting for dual-currency
        gap = params.gap_analysis
        if gap:
            adj_price = gap.get("adjusted_price")
            reported_cur = params.dcf_results.get("reported_currency", "")
            stock_cur = params.company_profile.get("currency", "")
            forex_rate = params.forex_rate
            if (adj_price and reported_cur and stock_cur
                    and reported_cur != stock_cur and forex_rate and forex_rate > 0):
                gap["adjusted_price_reporting"] = adj_price / forex_rate

        row_id = save_to_db(
            db_path=db_path,
            ticker=params.ticker,
            company_name=params.company_name,
            valuation_date=date.today().isoformat(),
            mode=params.mode,
            ai_engine=params.ai_engine,
            valuation_params=params.valuation_params,
            results=results,
            company_profile=params.company_profile,
            gap_analysis_result=gap,
            ai_result=params.ai_result,
            sensitivity_table=sens_table,
            wacc_sensitivity=wacc_sens,
            financial_data=financial_data,
            forex_rate=params.forex_rate,
            source='web',
            user_id=user_id,
        )
        return {"saved": True, "id": row_id}
    except Exception as e:
        return {"saved": False, "reason": str(e)}


# ── Export DCF to Excel ──────────────────────────────────────────────

@router.post("/export-excel")
def export_dcf_excel(params: DCFParams):
    """Generate DCF valuation Excel workbook and return as download.

    Re-runs the DCF calculation server-side and writes results to an
    Excel file using the existing template, returned as a streaming download.
    """
    import io as _io

    is_valid, err = validate_ticker(params.ticker)
    if not is_valid:
        raise HTTPException(status_code=400, detail=err)

    normalized = _normalize_ticker(params.ticker)

    # Fetch all needed data (same as run_dcf)
    financial_data = get_historical_financials(
        normalized, "annual", params.apikey, HISTORICAL_DATA_PERIODS_ANNUAL
    )
    if financial_data is None:
        raise HTTPException(status_code=404, detail=f"Financial data not found: {params.ticker}")

    profile = cached_get_profile(normalized, params.apikey)

    share_info = get_company_share_float(normalized, params.apikey, company_profile=profile)
    summary_df = financial_data["summary"]

    base_year_col = summary_df.columns[0]
    base_year_data = summary_df.iloc[:, 0].copy()
    base_year_data.name = base_year_col
    base_year_data["Outstanding Shares"] = share_info.get("outstandingShares", 0) or 0
    base_year_data["Average Tax Rate"] = financial_data["average_tax_rate"]
    base_year_data["Revenue Growth (%)"] = summary_df.iloc[
        summary_df.index.get_loc("Revenue Growth (%)"), 0
    ]
    base_year_data["Total Reinvestment"] = summary_df.iloc[
        summary_df.index.get_loc("Total Reinvestment"), 0
    ]

    # TTM detection
    ttm_quarter = financial_data.get("ttm_latest_quarter", "")
    ttm_end_date = financial_data.get("ttm_end_date", "")
    is_ttm = bool(ttm_quarter and ttm_end_date)
    fy_end_month = financial_data.get("fy_end_month", 12)
    base_year = int(str(base_year_col).replace('FY', ''))
    if is_ttm:
        _em = int(ttm_end_date[5:7])
        _ey = int(ttm_end_date[:4])
        forecast_year_1 = _ey if _em <= 6 else _ey + 1
    else:
        forecast_year_1 = base_year if fy_end_month <= 6 else base_year + 1
    ttm_label = f"{base_year_col}{ttm_quarter} TTM" if is_ttm else ""

    risk_free_rate = get_risk_free_rate(profile.get("country", "United States"))
    wacc_calc, total_erp, wacc_details = calculate_wacc(
        base_year_data, profile, params.apikey, verbose=False
    )
    tax_rate = params.tax_rate if params.tax_rate is not None else financial_data["average_tax_rate"] * 100
    wacc_val = params.wacc if params.wacc is not None else wacc_calc * 100
    if params.ronic_match_wacc:
        ronic = risk_free_rate + TERMINAL_RISK_PREMIUM
    else:
        ronic = risk_free_rate + TERMINAL_RISK_PREMIUM + TERMINAL_RONIC_PREMIUM

    raw_params = {
        "revenue_growth_1": params.revenue_growth_1,
        "revenue_growth_2": params.revenue_growth_2,
        "ebit_margin": params.ebit_margin,
        "convergence": params.convergence,
        "revenue_invested_capital_ratio_1": params.revenue_invested_capital_ratio_1,
        "revenue_invested_capital_ratio_2": params.revenue_invested_capital_ratio_2,
        "revenue_invested_capital_ratio_3": params.revenue_invested_capital_ratio_3,
        "tax_rate": tax_rate,
        "wacc": wacc_val,
        "ronic": ronic,
    }
    valuation_params = _build_valuation_params(
        raw_params, base_year, risk_free_rate, is_ttm, ttm_quarter, ttm_label,
        forecast_year_1=forecast_year_1, fy_end_month=fy_end_month
    )

    results = calculate_dcf(
        base_year_data, valuation_params, financial_data, share_info, profile
    )

    # WACC sensitivity
    wacc_results, wacc_base = wacc_sensitivity_analysis(
        base_year_data, valuation_params, financial_data, share_info, profile
    )

    # Sensitivity (growth × margin)
    sens_table = sensitivity_analysis(
        base_year_data, valuation_params, financial_data, share_info, profile
    )

    # Write to Excel in memory (self-contained, no template dependency)
    buf = _io.BytesIO()
    _build_dcf_excel(
        buf, base_year_data, financial_data, valuation_params,
        profile, results, total_erp, sens_table,
        (wacc_results, wacc_base),
    )
    buf.seek(0)

    company_name = profile.get("companyName", normalized).replace("/", "_")
    filename = f"{company_name} DCF {base_year}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_dcf_excel(buf, base_year_data, financial_data, valuation_params,
                     profile, results, total_erp, sens_table, wacc_sensitivity):
    """Build DCF Excel workbook: 3 sheets mirroring the DB record layout.

    Sheet 1 — DCF Valuation: forecast table + bridge + input parameters
    Sheet 2 — Sensitivity: growth×margin matrix + WACC sensitivity
    Sheet 3 — Historical Data: financial summary
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd

    wb = Workbook()
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_text = Font(bold=True, color="FFFFFF", size=11)
    section_font = Font(bold=True, size=11, color="2F5496")
    pct_fmt = '0.0%'
    amt_fmt = '#,##0'
    company_name = profile.get('companyName', 'N/A')
    reported_ccy = results.get('reported_currency', '')

    # ════════════════════════════════════════════════════════════
    # Sheet 1: DCF Valuation  (forecast + bridge + params)
    # ════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "DCF Valuation"
    row_n = 1
    ws.cell(row=row_n, column=1,
            value=f"{company_name} — DCF Valuation (in {reported_ccy}, millions)").font = hdr_font
    row_n += 2

    # ── Forecast Table ──
    dcf_table = results.get('dcf_table')
    if dcf_table is not None and isinstance(dcf_table, pd.DataFrame) and not dcf_table.empty:
        ws.cell(row=row_n, column=1, value="Forecast Table").font = section_font
        row_n += 1
        cols = list(dcf_table.columns)
        for ci, col_name in enumerate(cols, 1):
            c = ws.cell(row=row_n, column=ci, value=col_name)
            c.font = hdr_text
            c.fill = hdr_fill
        row_n += 1
        pct_rows_set = {'Revenue Growth Rate', 'EBIT Margin', 'Tax to EBIT', 'WACC', 'Discount Factor'}
        for _, frow in dcf_table.iterrows():
            for ci, col_name in enumerate(cols, 1):
                v = frow[col_name]
                if isinstance(v, (np.integer, np.floating)):
                    v = float(v)
                cell = ws.cell(row=row_n, column=ci, value=v)
                if isinstance(v, (int, float)) and col_name != 'Year':
                    cell.number_format = pct_fmt if col_name in pct_rows_set else amt_fmt
            row_n += 1
        # Auto-fit forecast columns
        for ci in range(1, len(cols) + 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 16

    # ── Valuation Bridge ──
    row_n += 2
    ws.cell(row=row_n, column=1, value="Valuation Bridge").font = section_font
    row_n += 1
    bridge_items = [
        ("PV (FCFF next 10 years)", results.get('pv_cf_next_10_years', 0)),
        ("PV (Terminal Value)", results.get('pv_terminal_value', 0)),
        ("Operating Value", (results.get('pv_cf_next_10_years', 0) or 0) + (results.get('pv_terminal_value', 0) or 0)),
        None,
        ("(+) Cash & Equivalents", results.get('cash', 0)),
        ("(+) Total Investments", results.get('total_investments', 0)),
        ("(-) Total Debt", results.get('total_debt', 0)),
        ("(-) Minority Interest", results.get('minority_interest', 0)),
        None,
        ("Equity Value", results.get('equity_value', 0)),
        ("Outstanding Shares (M)", results.get('outstanding_shares', 0)),
        ("Price per Share", results.get('price_per_share', 0)),
    ]
    for item in bridge_items:
        if item is None:
            row_n += 1
            continue
        label, val = item
        ws.cell(row=row_n, column=1, value=label)
        c = ws.cell(row=row_n, column=2, value=val)
        c.number_format = amt_fmt
        if label in ("Operating Value", "Equity Value", "Price per Share"):
            ws.cell(row=row_n, column=1).font = Font(bold=True)
            c.font = Font(bold=True)
        row_n += 1

    # ── Input Parameters ──
    row_n += 2
    ws.cell(row=row_n, column=1, value="Valuation Parameters").font = section_font
    row_n += 1
    param_items = [
        ("Base year", valuation_params['base_year'], None),
        ("Revenue growth Y1", valuation_params['revenue_growth_1'] / 100, pct_fmt),
        ("Revenue growth Y2-5 (CAGR)", valuation_params['revenue_growth_2'] / 100, pct_fmt),
        ("Risk-free rate (terminal growth)", valuation_params['risk_free_rate'], pct_fmt),
        ("Target EBIT margin", valuation_params['ebit_margin'] / 100, pct_fmt),
        ("Years to target margin", valuation_params['convergence'], None),
        ("Revenue/IC ratio (Y1-2)", valuation_params['revenue_invested_capital_ratio_1'], None),
        ("Revenue/IC ratio (Y3-5)", valuation_params['revenue_invested_capital_ratio_2'], None),
        ("Revenue/IC ratio (Y6-10)", valuation_params['revenue_invested_capital_ratio_3'], None),
        ("WACC (Y1-5)", valuation_params['wacc'] / 100, pct_fmt),
        ("Terminal WACC", valuation_params['risk_free_rate'] + TERMINAL_RISK_PREMIUM, pct_fmt),
        ("RONIC", valuation_params['ronic'], pct_fmt),
        ("Effective tax rate", valuation_params['tax_rate'] / 100, pct_fmt),
        None,
        ("Risk-free rate", valuation_params['risk_free_rate'], pct_fmt),
        ("Cost of debt", base_year_data.get('Cost of Debt (%)', 0) / 100, pct_fmt),
        ("Total equity risk premium", total_erp, pct_fmt),
        ("Beta", profile.get('beta', 1.0), None),
    ]
    for item in param_items:
        if item is None:
            row_n += 1
            continue
        label, val, fmt = item
        ws.cell(row=row_n, column=1, value=label)
        c = ws.cell(row=row_n, column=2, value=val)
        if fmt:
            c.number_format = fmt
        row_n += 1
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 16

    # ════════════════════════════════════════════════════════════
    # Sheet 2: Sensitivity Analysis
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Sensitivity")
    r2 = 1
    ws2.cell(row=r2, column=1, value="Sensitivity: Revenue Growth × EBIT Margin").font = hdr_font
    r2 += 2

    if hasattr(sens_table, 'values'):
        margins = list(sens_table.index) if hasattr(sens_table, 'index') else []
        growths = list(sens_table.columns) if hasattr(sens_table, 'columns') else []
        ws2.cell(row=r2, column=1, value="Growth \\ Margin").font = Font(bold=True, size=10)
        for j, m in enumerate(margins):
            c = ws2.cell(row=r2, column=2 + j,
                         value=m / 100 if isinstance(m, (int, float)) and abs(m) > 1 else m)
            c.number_format = pct_fmt
            c.font = Font(bold=True, size=10)
        for gi, g in enumerate(growths):
            row = r2 + 1 + gi
            c = ws2.cell(row=row, column=1,
                         value=g / 100 if isinstance(g, (int, float)) and abs(g) > 1 else g)
            c.number_format = pct_fmt
            c.font = Font(bold=True, size=10)
            for mi in range(len(margins)):
                val = sens_table.values[mi][gi] if hasattr(sens_table, 'values') else 0
                ws2.cell(row=row, column=2 + mi, value=val).number_format = amt_fmt
        r2 += len(growths) + 3

    if wacc_sensitivity:
        wacc_results, wacc_base = wacc_sensitivity
        ws2.cell(row=r2, column=1, value="WACC Sensitivity Analysis").font = hdr_font
        r2 += 1
        ws2.cell(row=r2, column=1, value="WACC")
        ws2.cell(row=r2 + 1, column=1, value="Price / Share")
        for j, (wv, price) in enumerate(wacc_results.items()):
            col = 2 + j
            c1 = ws2.cell(row=r2, column=col, value=wv / 100)
            c1.number_format = pct_fmt
            c2 = ws2.cell(row=r2 + 1, column=col, value=price)
            c2.number_format = amt_fmt
            if wv == wacc_base:
                c1.font = Font(bold=True)
                c2.font = Font(bold=True)

    ws2.column_dimensions['A'].width = 18
    for ci in range(2, 15):
        ws2.column_dimensions[chr(64 + ci)].width = 12

    # ════════════════════════════════════════════════════════════
    # Sheet 3: Historical Financial Data
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Historical Data")
    summary_df = financial_data['summary']
    for r in dataframe_to_rows(summary_df, index=True, header=True):
        ws3.append(r)
    AMOUNT_ROWS = {'Revenue', 'EBIT', '(+) Capital Expenditure', '(-) D&A',
                   '(+) ΔWorking Capital', 'Total Reinvestment',
                   '(+) Total Debt', '(+) Total Equity', 'Minority Interest',
                   '(-) Cash & Equivalents', '(-) Total Investments', 'Invested Capital'}
    RATIO_ROWS = {'Revenue Growth (%)', 'EBIT Growth (%)', 'EBIT Margin (%)', 'Tax Rate (%)',
                  'Revenue / IC', 'Debt to Assets (%)', 'Cost of Debt (%)',
                  'ROIC (%)', 'ROE (%)', 'Dividend Yield (%)', 'Payout Ratio (%)'}
    for row in ws3.iter_rows(min_row=2):
        label = row[0].value
        if label in AMOUNT_ROWS:
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
        elif label in RATIO_ROWS:
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0'
    for col_cells in ws3.columns:
        col_letter = col_cells[0].column_letter
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws3.column_dimensions[col_letter].width = min(max_len + 3, 30)

    wb.save(buf)
