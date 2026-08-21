# Copyright (c) 2025-2026 Alan He. Licensed under AGPL-3.0. See LICENSE.
"""Portfolio Tracker API — positions, cash, snapshots, KPI.

All data lives in a local SQLite database (PORTFOLIO_DB_PATH env var).
This router is only active when the database file exists, keeping the
public deployment unaffected.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import sqlite3
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from backend.routers.auth import get_current_user

logger = logging.getLogger("valuescope.portfolio.api")

router = APIRouter()

# ── Helpers ──────────────────────────────────────────────

def _get_db_path() -> str:
    from backend.services.portfolio_db import DB_PATH
    return DB_PATH


def _require_db() -> str:
    """Return DB_PATH, auto-creating if missing."""
    path = _get_db_path()
    if not os.path.exists(path):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        from backend.services.portfolio_db import init_db
        init_db(path)
        logger.info("Auto-created portfolio DB at %s", path)
    return path


def _query(path: str, sql: str, params=()):
    """Execute SQL and return list of dicts."""
    with sqlite3.connect(path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]


# ── Pydantic models ──────────────────────────────────────

class PositionIn(BaseModel):
    ticker: str
    name: str
    market: str
    broker: str
    quantity: float
    cost_price: float
    currency: str = "CNY"
    status: str = "open"


class CashIn(BaseModel):
    account: str
    currency: str
    balance: float


class AccountSettingIn(BaseModel):
    broker: str
    capital_mode: str = "cost"  # 'cost' or 'deposit'
    deposit_cny: float = 0
    deposit_fx: float = 1.0
    notes: str = ""
    cost_method: str = "diluted"  # 'diluted' (re-avg on sell) or 'average' (IBKR)
    hk_connect: bool = False  # 港股通: HK sale proceeds settle in CNY


class DepositRecordIn(BaseModel):
    broker: str
    amount_cny: float  # signed: deposit > 0, withdrawal < 0
    fx_rate: float = 1.0
    deposit_date: str = ""
    notes: str = ""
    currency: str = "CNY"
    amount: Optional[float] = None  # original-currency amount (defaults to amount_cny / fx_rate)
    update_cash: bool = False  # also move the matching cash balance


class MarginIn(BaseModel):
    broker: str
    category: str  # 'in_house' or 'off_exchange'
    currency: str = "CNY"
    amount: float


class ClosedTradeIn(BaseModel):
    ticker: str
    name: str
    market: str
    broker: str
    quantity: float
    buy_price: float
    sell_price: float
    realized_pnl: float
    realized_pnl_cny: float
    currency: str = "CNY"


# ── Endpoints ────────────────────────────────────────────

@router.get("/status")
def portfolio_status(user_id: Optional[str] = Depends(get_current_user)):
    """Check if portfolio feature is available.

    For logged-in users, auto-create the database so the onboarding
    flow (account setup, import templates) is shown instead of a
    technical 'not configured' error.
    """
    path = _get_db_path()
    if not os.path.exists(path) and user_id and user_id != "local":
        os.makedirs(os.path.dirname(path), exist_ok=True)
        from backend.services.portfolio_db import init_db
        init_db(path)
        logger.info("Auto-created portfolio DB for user %s at %s", user_id, path)
    exists = os.path.exists(path)
    return {"available": exists, "db_path": path if exists else None}


@router.get("/portfolios")
def list_portfolios():
    """Return available portfolios and which is currently active."""
    from backend.services.portfolio_db import PORTFOLIOS, DB_PATH
    items = [{"name": name, "active": os.path.abspath(path) == os.path.abspath(DB_PATH)}
             for name, path in PORTFOLIOS]
    if not items:
        items = [{"name": "Default", "active": True}]
    return items


@router.post("/portfolios/switch")
def switch_portfolio(name: str = Query(...)):
    """Switch the active portfolio DB by name."""
    from backend.services.portfolio_db import PORTFOLIOS, set_active_portfolio, init_db
    for pname, ppath in PORTFOLIOS:
        if pname == name:
            set_active_portfolio(ppath)
            init_db()
            return {"active": name}
    raise HTTPException(404, f"Portfolio '{name}' not found")


@router.get("/positions")
def list_positions(status: str = Query("open", pattern="^(open|closed|all)$"), user_id: str = Depends(get_current_user)):
    """List portfolio positions."""
    path = _require_db()
    if status == "all":
        sql = "SELECT * FROM positions WHERE user_id=? ORDER BY market, broker, name"
        return _query(path, sql, (user_id,))
    else:
        sql = "SELECT * FROM positions WHERE status=? AND user_id=? ORDER BY market, broker, name"
        return _query(path, sql, (status, user_id))


@router.post("/positions")
def upsert_position_api(pos: PositionIn, user_id: str = Depends(get_current_user)):
    """Create or update a position (upsert by ticker+broker)."""
    _require_db()
    from backend.services.portfolio_db import init_db, upsert_position, get_conn
    init_db()
    with get_conn() as conn:
        upsert_position(
            conn, ticker=pos.ticker, name=pos.name, market=pos.market,
            broker=pos.broker, quantity=pos.quantity,
            cost_price=pos.cost_price, currency=pos.currency, user_id=user_id,
        )
        conn.commit()
    return {"ok": True}


@router.delete("/positions/{ticker}/{broker}")
def delete_position_api(ticker: str, broker: str, user_id: str = Depends(get_current_user)):
    """Delete a position by ticker+broker."""
    path = _require_db()
    import sqlite3 as _sqlite3
    with _sqlite3.connect(path) as conn:
        cursor = conn.execute(
            "DELETE FROM positions WHERE ticker=? AND broker=? AND user_id=?", (ticker, broker, user_id)
        )
        conn.commit()
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Position not found")
    return {"ok": True}


@router.get("/closed-trades")
def list_closed_trades(user_id: str = Depends(get_current_user)):
    """List all closed (realized) trades."""
    path = _require_db()
    return _query(path, "SELECT * FROM closed_trades WHERE user_id=? ORDER BY market, abs(realized_pnl) DESC", (user_id,))


@router.post("/closed-trades")
def add_closed_trade(trade: ClosedTradeIn, user_id: str = Depends(get_current_user)):
    """Record a closed trade."""
    _require_db()
    import datetime as _dt
    from backend.services.portfolio_db import (
        init_db, insert_closed_trade, get_conn, compute_locked_ytd_cny,
    )
    init_db()
    with get_conn() as conn:
        # Lock this sale's YTD attribution now, while the baseline still
        # describes the lot being sold — immune to later baseline resets
        # (full close + re-buy) and FX drift.
        ytd_locked = compute_locked_ytd_cny(
            conn, ticker=trade.ticker, broker=trade.broker, market=trade.market,
            currency=trade.currency, quantity=trade.quantity,
            close_price=trade.sell_price, realized_pnl=trade.realized_pnl,
            realized_pnl_cny=trade.realized_pnl_cny, user_id=user_id,
        )
        insert_closed_trade(
            conn, ticker=trade.ticker, name=trade.name, market=trade.market,
            broker=trade.broker, currency=trade.currency,
            realized_pnl=trade.realized_pnl,
            realized_pnl_cny=trade.realized_pnl_cny,
            quantity=trade.quantity,
            cost_price=trade.buy_price, close_price=trade.sell_price,
            close_date=_dt.date.today().isoformat(),
            ytd_pnl_cny_locked=ytd_locked,
            user_id=user_id,
        )
        conn.commit()
    return {"ok": True}


@router.get("/cash")
def list_cash(user_id: str = Depends(get_current_user)):
    """List cash balances across all accounts."""
    path = _require_db()
    return _query(path, "SELECT * FROM cash_balances WHERE user_id=? ORDER BY account", (user_id,))


@router.post("/cash")
def update_cash(cash: CashIn, user_id: str = Depends(get_current_user)):
    """Update cash balance for an account."""
    _require_db()
    from backend.services.portfolio_db import init_db, upsert_cash, get_conn
    init_db()
    with get_conn() as conn:
        upsert_cash(conn, account=cash.account, currency=cash.currency, balance=cash.balance, user_id=user_id)
        conn.commit()
    return {"ok": True}


@router.delete("/cash/{account}/{currency}")
def delete_cash_api(account: str, currency: str, user_id: str = Depends(get_current_user)):
    """Delete a cash balance row by account+currency."""
    path = _require_db()
    import sqlite3 as _sqlite3
    with _sqlite3.connect(path) as conn:
        cursor = conn.execute(
            "DELETE FROM cash_balances WHERE account=? AND currency=? AND user_id=?",
            (account, currency, user_id))
        conn.commit()
    if cursor.rowcount == 0:
        raise HTTPException(404, "Cash row not found")
    return {"ok": True}


@router.get("/account-settings")
def list_account_settings(user_id: str = Depends(get_current_user)):
    """List all account capital settings."""
    _require_db()
    from backend.services.portfolio_db import init_db, get_account_settings, get_conn
    init_db()
    with get_conn() as conn:
        return get_account_settings(conn, user_id=user_id)


@router.post("/account-settings")
def upsert_account_setting_api(setting: AccountSettingIn, user_id: str = Depends(get_current_user)):
    """Create or update account capital settings."""
    _require_db()
    from backend.services.portfolio_db import init_db, upsert_account_setting, get_conn
    init_db()
    with get_conn() as conn:
        upsert_account_setting(
            conn, broker=setting.broker, capital_mode=setting.capital_mode,
            deposit_cny=setting.deposit_cny, deposit_fx=setting.deposit_fx,
            notes=setting.notes, cost_method=setting.cost_method, user_id=user_id,
            hk_connect=setting.hk_connect,
        )
        conn.commit()
    return {"ok": True}


@router.delete("/account-settings/{broker}")
def delete_account_setting_api(broker: str, user_id: str = Depends(get_current_user)):
    """Delete account capital settings (reverts to cost mode)."""
    _require_db()
    from backend.services.portfolio_db import init_db, delete_account_setting, get_conn
    init_db()
    with get_conn() as conn:
        affected = delete_account_setting(conn, broker, user_id=user_id)
        conn.commit()
        if affected == 0:
            raise HTTPException(status_code=404, detail="Account setting not found")
    return {"ok": True}


class MergeAccountRequest(BaseModel):
    source: str       # account to merge FROM (will be deleted)
    target: str       # account to merge INTO (will be kept)


@router.post("/merge-accounts")
def merge_accounts(req: MergeAccountRequest, user_id: str = Depends(get_current_user)):
    """Merge one account into another: rename all records from source to target, then delete source settings."""
    _require_db()
    from backend.services.portfolio_db import init_db, get_conn
    init_db()

    if req.source == req.target:
        raise HTTPException(status_code=400, detail="Source and target must be different")

    with get_conn() as conn:
        # Count affected records
        counts = {}
        for table, col in [("positions", "broker"), ("cash_balances", "account"),
                           ("closed_trades", "broker"), ("deposit_history", "broker")]:
            try:
                cur = conn.execute(f"SELECT COUNT(*) FROM {table} WHERE {col}=? AND user_id=?",
                                   (req.source, user_id))
                counts[table] = cur.fetchone()[0]
            except Exception:
                counts[table] = 0

        if sum(counts.values()) == 0:
            # No records to merge — just check if source account_settings exists
            cur = conn.execute("SELECT COUNT(*) FROM account_settings WHERE broker=? AND user_id=?",
                               (req.source, user_id))
            if cur.fetchone()[0] == 0:
                raise HTTPException(status_code=404, detail=f"Account '{req.source}' not found")

        # Rename all records from source to target
        for table, col in [("positions", "broker"), ("closed_trades", "broker"),
                           ("deposit_history", "broker")]:
            try:
                conn.execute(f"UPDATE {table} SET {col}=? WHERE {col}=? AND user_id=?",
                             (req.target, req.source, user_id))
            except Exception:
                pass

        # Cash balances: merge by summing if target already has same currency entry
        try:
            source_cash = conn.execute(
                "SELECT currency, balance FROM cash_balances WHERE account=? AND user_id=?",
                (req.source, user_id)
            ).fetchall()
            for currency, balance in source_cash:
                existing = conn.execute(
                    "SELECT balance FROM cash_balances WHERE account=? AND currency=? AND user_id=?",
                    (req.target, currency, user_id)
                ).fetchone()
                if existing:
                    conn.execute(
                        "UPDATE cash_balances SET balance=?, updated_at=datetime('now','localtime') "
                        "WHERE account=? AND currency=? AND user_id=?",
                        (existing[0] + balance, req.target, currency, user_id)
                    )
                else:
                    conn.execute(
                        "UPDATE cash_balances SET account=? WHERE account=? AND currency=? AND user_id=?",
                        (req.target, req.source, currency, user_id)
                    )
            # Delete any remaining source cash rows
            conn.execute("DELETE FROM cash_balances WHERE account=? AND user_id=?",
                         (req.source, user_id))
        except Exception:
            pass

        # Delete source account_settings
        conn.execute("DELETE FROM account_settings WHERE broker=? AND user_id=?",
                     (req.source, user_id))

        conn.commit()

    return {
        "ok": True,
        "merged": counts,
        "message": f"Merged '{req.source}' into '{req.target}'",
    }


@router.get("/flows")
def list_all_flows(limit: int = Query(200, ge=1, le=1000),
                   user_id: str = Depends(get_current_user)):
    """All cash-flow records across brokers, newest first (Flows tab)."""
    path = _require_db()
    return _query(path, """
        SELECT * FROM deposit_history WHERE user_id=?
        ORDER BY COALESCE(deposit_date, substr(created_at, 1, 10)) DESC, id DESC
        LIMIT ?
    """, (user_id, limit))


@router.get("/deposit-history/{broker}")
def list_deposit_history(broker: str, user_id: str = Depends(get_current_user)):
    """List deposit history records for a broker."""
    _require_db()
    from backend.services.portfolio_db import init_db, get_deposit_history, get_conn
    init_db()
    with get_conn() as conn:
        return get_deposit_history(conn, broker, user_id=user_id)


@router.post("/deposit-history")
def add_deposit_record_api(data: DepositRecordIn, user_id: str = Depends(get_current_user)):
    """Add a deposit record and auto-recalculate totals."""
    _require_db()
    from backend.services.portfolio_db import init_db, add_deposit_record, get_conn
    init_db()
    with get_conn() as conn:
        add_deposit_record(conn, data.broker, data.amount_cny,
                           data.fx_rate, data.deposit_date, data.notes, user_id=user_id,
                           currency=data.currency, amount=data.amount,
                           update_cash=data.update_cash)
        conn.commit()
    return {"ok": True}


@router.delete("/deposit-history/{record_id}")
def delete_deposit_record_api(record_id: int,
                              force: bool = Query(False),
                              user_id: str = Depends(get_current_user)):
    """Delete a cash-flow record and recalculate totals.

    A flow the daily snapshot has already folded into units is refused with
    409 — deleting it would move NAV with no matching units change, i.e. a
    phantom gain/loss on the TWR curve. Book a reverse flow instead, or pass
    force=true for genuine data repair.
    """
    _require_db()
    from backend.services.portfolio_db import init_db, delete_deposit_record, get_conn
    init_db()
    with get_conn() as conn:
        res = delete_deposit_record(conn, record_id, user_id=user_id, force=force)
        conn.commit()
        if res["reason"] == "not_found":
            raise HTTPException(status_code=404, detail="Record not found")
        if not res["deleted"]:
            raise HTTPException(status_code=409, detail={
                "error_code": "flow_already_unitized",
                "message": (f"这笔流水已在 {res['unitized_on']} 的快照里折算成份额，"
                            "删除会让单位净值凭空跳动。正确做法是记一笔等额反向流水；"
                            "确需修数据可强制删除（现金不会回滚）。"),
            })
    return {"ok": True, **res}


@router.get("/margin")
def list_margin(user_id: str = Depends(get_current_user)):
    """List margin/leverage balances."""
    path = _require_db()
    return _query(path, "SELECT * FROM margin_balances WHERE user_id=? ORDER BY broker", (user_id,))


@router.post("/margin")
def update_margin(m: MarginIn, user_id: str = Depends(get_current_user)):
    """Update a margin/leverage balance."""
    _require_db()
    from backend.services.portfolio_db import init_db, upsert_margin, get_conn
    init_db()
    with get_conn() as conn:
        upsert_margin(conn, broker=m.broker, category=m.category,
                      currency=m.currency, amount=m.amount, user_id=user_id)
        conn.commit()
    return {"ok": True}


@router.get("/fx-rates")
def get_fx_rates():
    """Get live FX rates (with DB fallback)."""
    path = _require_db()
    from backend.services.portfolio_prices import get_fx_rates as _get_fx
    return _get_fx(db_path=path)


@router.get("/prices/{ticker}")
def get_price(ticker: str):
    """Fetch live price for a single ticker."""
    from backend.services.portfolio_prices import fetch_price, get_previous_close
    price, currency = fetch_price(ticker)
    prev_close = get_previous_close(ticker)
    if price is None:
        raise HTTPException(status_code=404, detail=f"No price for {ticker}")
    return {"ticker": ticker, "price": price, "currency": currency, "prev_close": prev_close}


@router.get("/holdings")
def get_enriched_holdings(user_id: str = Depends(get_current_user)):
    """Get positions enriched with live prices, P&L, daily/YTD returns.

    This is the main data endpoint for the portfolio dashboard — combines
    positions + prices + FX + YTD baselines into a single response.
    """
    path = _require_db()
    from backend.services.portfolio_db import init_db, get_ytd_baselines, get_conn, get_dcf_valuations, compute_capital
    from backend.services.portfolio_prices import (
        fetch_price, get_previous_close, get_fx_rates as _get_fx,
        refresh_all_prices, get_price_delay_minutes, get_price_as_of,
        get_price_source, get_price_session, is_price_stale,
    )
    from backend.services.ytd_calc import held_ytd
    import pandas as pd

    init_db()

    # Load positions
    positions = _query(path, "SELECT * FROM positions WHERE status='open' AND user_id=? ORDER BY market, broker, name", (user_id,))
    if not positions:
        return {"holdings": [], "fx": {"CNY": 1.0}, "summary": {}}

    # Per-broker cost convention. 'average' brokers (e.g. Interactive Brokers)
    # keep the true weighted-average cost on a partial sell — realized P&L is
    # booked in closed_trades — so their stocks get fund-style YTD attribution.
    # 'diluted' (default) re-averages cost on sell to absorb realized gains.
    broker_cost_method = {
        r['broker']: (r['cost_method'] or 'diluted')
        for r in _query(path, "SELECT broker, cost_method FROM account_settings WHERE user_id=?", (user_id,))
    }

    # Load industry cache — backfill any tickers missing from cache
    industry_rows = _query(path, "SELECT ticker, sector, industry FROM industry_cache")
    industry_map = {r['ticker']: {'sector': r['sector'] or '', 'industry': r['industry'] or ''} for r in industry_rows}
    all_tickers = [p['ticker'] for p in positions]
    missing_tickers = [t for t in all_tickers if t not in industry_map or (not industry_map[t].get('sector') and not industry_map[t].get('industry'))]
    if missing_tickers:
        from backend.services.portfolio_fmp import fetch_all_industries
        fetched = fetch_all_industries(missing_tickers, db_path=path)
        for t, (sector, industry) in fetched.items():
            industry_map[t] = {'sector': sector or '', 'industry': industry or ''}

    # Load DCF valuations from ValueScope DB
    dcf_map = get_dcf_valuations()

    # Get FX rates
    fx = _get_fx(db_path=path)

    # Prefetch all prices in parallel
    refresh_all_prices(db_path=path)

    # Load YTD baselines — auto-create for new positions missing a baseline
    import datetime
    current_year = datetime.date.today().year
    with get_conn() as conn:
        ytd_baselines = get_ytd_baselines(conn, current_year, user_id=user_id)

        # Backfill baselines for positions added after the yearly snapshot
        missing_baseline = [
            p for p in positions
            if (p['ticker'], p['broker']) not in ytd_baselines
        ]
        if missing_baseline:
            from backend.services.portfolio_db import record_ytd_baselines
            # Two distinct cases share this backfill:
            # - year rollover (no baselines for the year at all): baseline =
            #   latest close (≈ year-end close), same rule as the snapshot
            #   job. Using cost here would silently turn the whole year's
            #   attribution into "since cost" — and INSERT OR IGNORE means
            #   whoever writes first wins permanently.
            # - a position opened mid-year: YTD starts from its cost.
            year_rollover = not ytd_baselines
            ticker_data = {}
            for p in missing_baseline:
                if year_rollover:
                    px, _ = fetch_price(p['ticker'])
                    base = px if px is not None else p['cost_price']
                else:
                    base = p['cost_price']
                ticker_data[(p['ticker'], p['broker'])] = (
                    base, p['currency'], p['quantity'], p['cost_price']
                )
            today_str = datetime.date.today().isoformat()
            record_ytd_baselines(conn, current_year, ticker_data, today_str, user_id=user_id)
            conn.commit()
            # Reload after backfill
            ytd_baselines = get_ytd_baselines(conn, current_year, user_id=user_id)

    # Enrich each position
    holdings = []
    total_equity_cny = 0
    total_cost_cny = 0
    total_pnl_cny = 0
    total_daily_pnl_cny = 0
    total_ytd_pnl_cny = 0

    # Lifetime per-position extras for the Total Return column:
    # realized P&L (closed_trades) + net dividends (dividend_log, Flex-fed).
    # Convention-safe by construction: diluted partial sells fold gains into
    # cost (no closed_trade row), and domestic dividends fold into diluted
    # cost (never in the ledger) — so nothing is double counted.
    realized_map: dict = {}
    for r in _query(path, "SELECT ticker, broker, currency, realized_pnl, realized_pnl_cny "
                          "FROM closed_trades WHERE user_id=? AND ticker IS NOT NULL", (user_id,)):
        v = r['realized_pnl_cny']
        if v is None:
            v = (r['realized_pnl'] or 0) * fx.get(r['currency'], 1.0)
        k = (r['ticker'], r['broker'])
        realized_map[k] = realized_map.get(k, 0.0) + v
    div_map: dict = {}  # (ticker, account) -> net CNY
    try:
        for r in _query(path, "SELECT ticker, account, currency, amount FROM dividend_log "
                              "WHERE user_id=? AND ticker != ''", (user_id,)):
            k = (r['ticker'], r['account'] or '')
            div_map[k] = div_map.get(k, 0.0) + r['amount'] * fx.get(r['currency'], 1.0)
    except Exception:
        pass  # ledger table appears with the first Flex sync

    for pos in positions:
        ticker = pos['ticker']
        qty = pos['quantity']
        cost = pos['cost_price']
        currency = pos['currency']
        rate = fx.get(currency, 1.0)

        # Live price
        price, _ = fetch_price(ticker)
        price_missing = price is None
        price_stale = price_missing or is_price_stale(ticker)
        price_date = get_price_as_of(ticker)
        price_source = get_price_source(ticker)
        if price is None:
            price = cost

        mv = qty * price
        cost_total = qty * cost
        pnl = mv - cost_total
        pnl_pct = (pnl / cost_total * 100) if cost_total != 0 else 0
        mv_cny = mv * rate
        pnl_cny = pnl * rate

        # Daily P&L
        # For positions created today, use cost as base instead of previous close
        import datetime as _dt
        _today_str = _dt.date.today().isoformat()
        _created_today = pos.get('created_at', '')[:10] == _today_str

        # Non-trading day handling is done at the data source level:
        # - Eastmoney: f86 timestamp check → prev_close = price when data is stale
        # - yfinance (non-USD): weekend + newest-bar-date checks (covers
        #   exchange holidays, e.g. JP 海の日) → prev_close = price when the
        #   exchange has no session today
        # - yfinance (USD extended): prev_close = last regular close; live
        #   pre-market (marketState PRE) counts as today's move, weekends and
        #   holidays (CLOSED) are zeroed
        # - Fund NAV: nav_date check → prev_nav = nav when stale
        # When prev_close == price, daily_pnl naturally = 0.

        prev_close = get_previous_close(ticker)
        if _created_today and not price_stale:
            # New position today: daily P&L = current price vs cost
            daily_pnl = (price - cost) * qty
            daily_pnl_pct = (price / cost - 1) * 100 if cost > 0 else 0
            daily_pnl_cny = daily_pnl * rate
        elif prev_close and prev_close > 0 and not price_stale:
            daily_pnl = (price - prev_close) * qty
            daily_pnl_pct = (price / prev_close - 1) * 100
            daily_pnl_cny = daily_pnl * rate
        else:
            daily_pnl = daily_pnl_pct = daily_pnl_cny = None

        # YTD P&L for held shares only. Sold shares' YTD contribution is captured
        # in closed_trades (aggregated below).
        #
        # Stocks use the cost-adjustment convention: cost_price is re-averaged
        # after each partial sell to absorb realized gains, so `pnl - baseline_unrealized`
        # gives correct YTD.
        #
        # Average-cost instruments don't adjust cost on a partial sell: funds
        # (market='基金', industry standard) and stocks on 'average' brokers like
        # Interactive Brokers. For a partial sell the held YTD is just price drift
        # on remaining units: `(price - bp) × current_qty`. The sold units' realized
        # YTD contribution is locked at sell-time in a closed_trade and picked up by
        # the partial-close branch of the closed_trades aggregation below.
        # Adds (qty > b_qty) still use the standard formula since cost is then a
        # weighted average and the algebra works out.
        uses_avg_cost = (
            pos.get('market') == '基金'
            or broker_cost_method.get(pos.get('broker'), 'diluted') == 'average'
        )
        ytd_pnl = ytd_pnl_pct = ytd_pnl_cny = None
        bd = None
        if ytd_baselines:
            bd = (ytd_baselines.get((ticker, pos['broker']))
                  or ytd_baselines.get((ticker, '')))  # legacy broker-less row
        ytd_pnl = held_ytd(price, cost, qty, bd, uses_avg_cost)
        if ytd_pnl is not None:
            ytd_pnl_pct = (ytd_pnl / cost_total * 100) if cost_total != 0 else 0
            ytd_pnl_cny = ytd_pnl * rate
        # else: no baseline — ytd_pnl stays None (KPI YTD uses snapshot-based calculation)

        total_equity_cny += mv_cny
        total_cost_cny += cost_total * rate
        total_pnl_cny += pnl_cny
        if daily_pnl_cny is not None:
            total_daily_pnl_cny += daily_pnl_cny
        if ytd_pnl_cny is not None:
            total_ytd_pnl_cny += ytd_pnl_cny

        # Industry
        ind = industry_map.get(ticker, {})
        # DCF
        dcf_info = dcf_map.get(ticker)
        dcf_price = dcf_info['dcf_price'] if dcf_info else None
        mos_pct = None
        if dcf_price and price and dcf_price > 0:
            mos_pct = round((dcf_price - price) / dcf_price * 100, 2)

        # Total return = unrealized + lifetime realized + net dividends
        # (attribution's 累计 tab uses the same three-ledger arithmetic)
        realized_cny = realized_map.get((ticker, pos['broker']), 0.0)
        dividends_cny = sum(
            v for (dtk, acct), v in div_map.items()
            if dtk == ticker and (not acct or acct in (pos['broker'] or '')))
        total_return_cny = pnl_cny + realized_cny + dividends_cny
        cost_cny = cost_total * rate
        total_return_pct = (total_return_cny / cost_cny * 100) if cost_cny > 0 else None

        holdings.append({
            **pos,
            # Keep internal arithmetic backwards-compatible when no valuation
            # has ever succeeded, but never expose cost basis as a market price.
            "price": None if price_missing else price,
            "price_stale": price_stale,
            "price_date": price_date,
            "price_source": price_source,
            # 'extended'/'night' when the US quote came from outside the
            # regular session, so the UI can say which book it is.
            "price_session": get_price_session(ticker),
            "market_value": mv,
            "market_value_cny": mv_cny,
            "cost_total": cost_total,
            "pnl": pnl,
            "pnl_pct": round(pnl_pct, 2),
            "pnl_cny": pnl_cny,
            "realized_cny": round(realized_cny, 2),
            "dividends_cny": round(dividends_cny, 2),
            "total_return_cny": round(total_return_cny, 2),
            "total_return_pct": round(total_return_pct, 2) if total_return_pct is not None else None,
            "daily_pnl": daily_pnl,
            "daily_pnl_pct": round(daily_pnl_pct, 2) if daily_pnl_pct is not None else None,
            "daily_pnl_cny": daily_pnl_cny,
            # Minutes this quote lags the exchange (0 = real time). Tokyo has
            # no free real-time feed, so those stay delayed and the UI says so.
            "price_delay_min": get_price_delay_minutes(ticker),
            "ytd_pnl": ytd_pnl,
            "ytd_pnl_pct": round(ytd_pnl_pct, 2) if ytd_pnl_pct is not None else None,
            "ytd_pnl_cny": ytd_pnl_cny,
            "sector": ind.get('sector', ''),
            "industry": ind.get('industry', ''),
            "dcf_price": dcf_price,
            "mos_pct": mos_pct,
            "weight": 0,  # filled below
        })

    # Calculate weights
    for h in holdings:
        h["weight"] = round(h["market_value_cny"] / total_equity_cny * 100, 2) if total_equity_cny > 0 else 0

    # Cash summary — auto-create cash rows for each currency a broker holds positions in
    cash_rows = _query(path, "SELECT * FROM cash_balances WHERE user_id=? ORDER BY account", (user_id,))
    existing_pairs = {(r['account'], r['currency']) for r in cash_rows}
    # Collect all (broker, currency) pairs from positions
    needed_pairs = set()
    for pos in positions:
        needed_pairs.add((pos['broker'], pos['currency']))
    missing_pairs = needed_pairs - existing_pairs
    if missing_pairs:
        with get_conn() as conn:
            for broker, currency in sorted(missing_pairs):
                conn.execute("""
                    INSERT INTO cash_balances (account, currency, balance, updated_at, user_id)
                    VALUES (?, ?, 0, datetime('now','localtime'), ?)
                """, (broker, currency, user_id))
            conn.commit()
        # Re-fetch after insert
        cash_rows = _query(path, "SELECT * FROM cash_balances WHERE user_id=? ORDER BY account", (user_id,))
    # IBKR-style: negative balances are margin loans, counted as leverage
    _cash_vals = [r['balance'] * fx.get(r['currency'], 1.0) for r in cash_rows]
    cash_cny = sum(v for v in _cash_vals if v >= 0)
    neg_cash_cny = sum(-v for v in _cash_vals if v < 0)

    # Margin / leverage (legacy in_house rows + negative cash + off-exchange)
    margin_rows = _query(path, "SELECT * FROM margin_balances WHERE user_id=?", (user_id,))
    leverage_cny = neg_cash_cny + sum(r['amount'] * fx.get(r['currency'], 1.0) for r in margin_rows)

    total_assets = total_equity_cny + cash_cny
    net_assets = total_assets - leverage_cny

    # Compute capital using the proper deposit-mode formula
    with get_conn() as conn:
        capital = compute_capital(conn, fx, user_id=user_id)
    total_pnl_capital = net_assets - capital  # Total P&L = Net Assets - Capital

    # YTD realized P&L from closed trades (by market).
    # For fully-closed tickers: aggregate Σ realized_pnl across all this year's trades
    # for that ticker and subtract baseline_unrealized once. Robust to share-count
    # changes, partial-sell + cost-basis adjustments, and mid-year buy-then-sell.
    # For partial-closed tickers (still in positions): contribute the YTD-attributable
    # slice of the realized gain, i.e. (close_price − baseline_price) × qty × fx.
    # This locks the sold portion's YTD at sell-time (won't drift with current price)
    # and combines correctly with the fund's held formula (price - bp) × current_qty.
    # Falls back to stored realized_pnl_cny when close_price/qty are missing
    # (legacy records, e.g. Excel-imported partials).
    import datetime as _dt
    from collections import defaultdict
    ytd_start = f"{_dt.date.today().year}-01-01"
    open_pairs = {(p['ticker'], p['broker']) for p in positions}
    ytd_realized_by_market: dict[str, float] = {}
    ytd_realized_total = 0.0

    def _baseline_for(tk: str, brk: str):
        return ytd_baselines.get((tk, brk)) or ytd_baselines.get((tk, ''))

    fc_groups: dict[tuple, dict] = defaultdict(
        lambda: {'realized_native': 0.0, 'market': 'Other', 'currency': 'CNY',
                 'locked_sum': 0.0, 'locked_all': True})

    for row in _query(path,
            "SELECT market, broker, ticker, currency, close_price, quantity, "
            "COALESCE(realized_pnl, 0) AS rpn, "
            "COALESCE(realized_pnl_cny, 0) AS rpl, "
            "ytd_pnl_cny_locked AS locked "
            "FROM closed_trades WHERE close_date >= ? AND user_id=?", (ytd_start, user_id)):
        ticker = row['ticker']
        mkt = row['market'] or 'Other'
        if ticker and (ticker, row['broker']) not in open_pairs:
            g = fc_groups[(ticker, row['broker'])]
            g['realized_native'] += row['rpn']
            g['market'] = mkt
            g['currency'] = row['currency']
            if row['locked'] is not None:
                g['locked_sum'] += row['locked']
            else:
                g['locked_all'] = False
        else:
            # Partial close: sold shares' YTD, locked at sell time.
            # Legacy rows (no locked value) fall back to computing from the
            # live baseline; last resort is the stored realized CNY.
            if row['locked'] is not None:
                contribution = row['locked']
            else:
                baseline = _baseline_for(ticker, row['broker']) if ticker else None
                close_price = row['close_price']
                qty_sold = row['quantity']
                if (baseline is not None and close_price is not None
                        and qty_sold is not None):
                    rate = fx.get(row['currency'], 1.0)
                    contribution = (close_price - baseline['price']) * qty_sold * rate
                else:
                    contribution = row['rpl']
            ytd_realized_by_market[mkt] = ytd_realized_by_market.get(mkt, 0) + contribution
            ytd_realized_total += contribution

    # Fully-closed (ticker, broker) groups. When every trade carries a locked
    # value, sum those (each sale locked at its own time). Otherwise use the
    # aggregate formula Σ realized − baseline_unrealized, robust to partial
    # sells with cost re-averaging and mid-year buy-then-sell.
    for (ticker, brk), g in fc_groups.items():
        rate = fx.get(g['currency'], 1.0)
        if g['locked_all']:
            contribution = g['locked_sum']
        else:
            baseline = _baseline_for(ticker, brk)
            if (baseline is not None
                    and baseline.get('cost_price') is not None
                    and baseline.get('quantity') is not None):
                baseline_unrealized = (baseline['price'] - baseline['cost_price']) * baseline['quantity']
                contribution = (g['realized_native'] - baseline_unrealized) * rate
            else:
                contribution = g['realized_native'] * rate
        ytd_realized_by_market[g['market']] = ytd_realized_by_market.get(g['market'], 0) + contribution
        ytd_realized_total += contribution

    total_ytd_pnl_cny += ytd_realized_total

    # Latest TWR unit NAV (official, from snapshots) + a live intraday
    # estimate: (live NAV − flows since the snapshot) / latest units.
    # The official series only advances at the daily snapshot; the estimate
    # is display-level convenience, like a fund's 盘中估值.
    unit_nav = None
    unit_nav_date = None
    unit_nav_est = None
    with get_conn() as conn:
        _r = conn.execute(
            "SELECT unit_nav, units, date FROM daily_snapshots "
            "WHERE user_id=? AND unit_nav IS NOT NULL AND units > 0 "
            "ORDER BY date DESC LIMIT 1",
            (user_id,)).fetchone()
        if _r:
            from backend.services.portfolio_db import pending_flows
            unit_nav, _units, unit_nav_date = _r[0], _r[1], _r[2]
            # Flows still pending unitization (recorded after the snapshot,
            # or backdated) are already out of live NAV but not yet out of
            # units — add them back or a withdrawal reads as a loss
            _flows = sum(a for _i, a in pending_flows(conn, user_id))
            unit_nav_est = (net_assets - _flows) / _units

    # YTD money-weighted return (Modified Dietz): P&L over time-weighted
    # invested capital — the professional "how much did MY money earn"
    # counterpart to the hero's TWR. Denominator = year-start NAV + each
    # flow weighted by its remaining fraction of the period.
    ytd_mwr = None
    ytd_mwr_start = None
    with get_conn() as conn:
        import datetime as _dt
        _jan1 = f"{_dt.date.today().year}-01-01"
        # Start at the first UNITIZED snapshot of the year: from T0 onward
        # every external flow (incl. living-expense withdrawals) is recorded
        # in deposit_history, so Dietz is clean. Earlier history has
        # unrecorded consumption outflows that would read as losses.
        _r0 = conn.execute(
            "SELECT date, net_assets, created_at FROM daily_snapshots "
            "WHERE user_id=? AND date>=? AND net_assets IS NOT NULL "
            "AND units IS NOT NULL AND units > 0 ORDER BY date ASC LIMIT 1",
            (user_id, _jan1)).fetchone()
        if _r0 and _r0[1]:
            _d0 = _dt.date.fromisoformat(_r0[0])
            _nav0 = _r0[1]
            _today = _dt.date.today()
            _total = max((_today - _d0).days, 1)
            # Flows after the start snapshot — including same-day ones
            # entered after it was taken (its NAV predates them)
            _flows = conn.execute(
                "SELECT deposit_date, amount_cny FROM deposit_history "
                "WHERE user_id=? AND deposit_date IS NOT NULL AND deposit_date != '' "
                "AND (deposit_date > ? OR (deposit_date = ? AND created_at > ?))",
                (user_id, _r0[0], _r0[0], _r0[2])).fetchall()
            _fsum = sum(f[1] for f in _flows)
            _wsum = 0.0
            for _fd, _amt in _flows:
                try:
                    _w = max(0.0, min(1.0, (_today - _dt.date.fromisoformat(_fd)).days / _total))
                except ValueError:
                    _w = 0.5
                _wsum += _amt * _w
            _denom = _nav0 + _wsum
            if _denom > 0:
                ytd_mwr = (net_assets - _nav0 - _fsum) / _denom * 100
                ytd_mwr_start = _r0[0]

    summary = {
        "unit_nav": round(unit_nav, 4) if unit_nav else None,
        "unit_nav_date": unit_nav_date,
        "unit_nav_est": round(unit_nav_est, 4) if unit_nav_est else None,
        "ytd_mwr": round(ytd_mwr, 2) if ytd_mwr is not None else None,
        "ytd_mwr_start": ytd_mwr_start,
        "equity_cny": round(total_equity_cny, 2),
        "cash_cny": round(cash_cny, 2),
        "leverage_cny": round(leverage_cny, 2),
        "total_assets": round(total_assets, 2),
        "net_assets": round(net_assets, 2),
        "capital": round(capital, 2),
        "total_pnl_cny": round(total_pnl_cny, 2),
        "total_pnl_capital": round(total_pnl_capital, 2),
        "total_cost_cny": round(total_cost_cny, 2),
        "total_pnl_pct": round(total_pnl_cny / total_cost_cny * 100, 2) if total_cost_cny != 0 else 0,
        "daily_pnl_cny": round(total_daily_pnl_cny, 2),
        "ytd_pnl_cny": round(total_ytd_pnl_cny, 2),
    }

    return {
        "holdings": holdings, "fx": fx, "cash": cash_rows, "summary": summary,
        "ytd_realized_by_market": {k: round(v, 2) for k, v in ytd_realized_by_market.items()},
    }


@router.get("/snapshots")
def list_snapshots(limit: int = Query(90, ge=1, le=365), user_id: str = Depends(get_current_user)):
    """Get daily snapshots (most recent first)."""
    path = _require_db()
    return _query(path, "SELECT * FROM daily_snapshots WHERE user_id=? ORDER BY date DESC LIMIT ?", (user_id, limit))


@router.get("/nav-history")
def get_nav_history(user_id: str = Depends(get_current_user)):
    """Get NAV history for performance charting."""
    path = _require_db()
    return _query(path, "SELECT * FROM nav_history WHERE user_id=? ORDER BY date", (user_id,))


@router.get("/benchmarks")
def get_benchmarks(start: str = "2024-01-01"):
    """Benchmark series in CNY terms: {name: [{date, close}, ...]}.

    Correctness measures (each one fixed an observed distortion):
    - closes selected by explicit (Close, ticker) label — concurrent
      yfinance downloads can return another symbol's data (see _FX_SANE)
    - S&P 500 / Hang Seng converted to CNY — the portfolio TWR is
      CNY-denominated; a USD-terms index hides the FX component (~1.7pp
      YTD 2026) and overstates the alternative
    - Yahoo's 000300.SS feed lags by days; when stale vs the other
      benchmarks it falls back to the 510300 ETF as a CSI300 proxy
      (rebased returns are what the chart uses, so the proxy is fine)
    - the fetch window starts 10 days before `start` so the frontend can
      rebase at the last close on/before the portfolio's first snapshot
      (snapshots exist on Saturdays; indices don't)
    - cached 1h — this endpoint used to re-download three series per view
    """
    try:
        import yfinance as yf
        import pandas as pd
    except ImportError:
        return {}
    import bisect
    import datetime as _dt
    from backend import persistent_cache as pc

    cache_key = f"benchmarks_cny:{start}"
    cached = pc.get(cache_key)
    if cached is not None:
        return cached

    fetch_start = (_dt.date.fromisoformat(start) - _dt.timedelta(days=10)).isoformat()
    end = (pd.Timestamp.now() + pd.DateOffset(days=1)).strftime("%Y-%m-%d")

    def dl(ticker: str):
        hist = yf.download(ticker, start=fetch_start, end=end,
                           progress=False, auto_adjust=True)
        if hist is None or hist.empty:
            return None
        close = (hist[("Close", ticker)] if isinstance(hist.columns, pd.MultiIndex)
                 else hist["Close"]).dropna()
        return close if not close.empty else None

    expected = (("CSI 300", "000300.SS", "CNY"),
                ("S&P 500", "^GSPC", "USD"),
                ("Nasdaq 100", "^NDX", "USD"),
                ("Hang Seng", "^HSI", "HKD"))
    series: dict[str, tuple] = {}  # name -> (close_series, ccy)
    for name, ticker, ccy in expected:
        for _attempt in range(2):  # one retry — Yahoo flakes transiently
            try:
                s = dl(ticker)
                if s is not None:
                    series[name] = (s, ccy)
                    break
            except Exception:
                pass

    # ETF proxy fallback for the US indices (same idea as 510300 for CSI300:
    # the chart uses rebased returns, so a tracking ETF is fine)
    for name, proxy in (("Nasdaq 100", "QQQ"), ("S&P 500", "SPY")):
        if name not in series:
            try:
                s = dl(proxy)
                if s is not None:
                    series[name] = (s, "USD")
            except Exception:
                pass

    # Stale-feed fallback: if CSI300 trails the freshest benchmark by more
    # than 4 days, use the 510300 ETF proxy instead
    try:
        if series:
            latest = max(s.index[-1] for s, _ in series.values())
            csi = series.get("CSI 300")
            if csi is None or (latest - csi[0].index[-1]).days > 4:
                proxy = dl("510300.SS")
                if proxy is not None:
                    series["CSI 300"] = (proxy, "CNY")
    except Exception:
        pass

    results: dict[str, list] = {}
    for name, (s, ccy) in series.items():
        rates = _fx_history(ccy, fetch_start) if ccy != "CNY" else []
        rd = [r[0] for r in rates]
        rv = [r[1] for r in rates]
        rows = []
        for ts, close in zip(s.index, s):
            d = pd.to_datetime(ts).strftime("%Y-%m-%d")
            v = float(close)
            if ccy != "CNY":
                if not rd:
                    rows = []
                    break  # no sane FX series — drop rather than mislead
                i = bisect.bisect_right(rd, d) - 1
                if i < 0:
                    continue
                v *= rv[i]
            rows.append({"date": d, "close": v})
        if rows:
            results[name] = rows
    if results:
        # A series that failed every fetch would otherwise vanish from the
        # chart for the whole cache TTL — backfill it from the last complete
        # response, and keep partial results cached only briefly so the gap
        # heals on the next request instead of an hour later.
        lastgood_key = f"benchmarks_cny_lastgood:{start}"
        prev = pc.get(lastgood_key) or {}
        for name, _, _ in expected:
            if name not in results and name in prev:
                results[name] = prev[name]
        pc.put(lastgood_key, results, ttl=7 * 86400)
        complete = all(name in results for name, _, _ in expected)
        # 10 min, not 1h: the hero/perf charts now show a live intraday
        # portfolio point, so a stale benchmark would skew beating/trailing
        pc.put(cache_key, results, ttl=600 if complete else 300)
    return results


_MARKET_CCY = {"A股": "CNY", "基金": "CNY", "B股": "HKD", "港股": "HKD",
               "日股": "JPY", "美股": "USD"}


# Plausible {ccy}CNY bounds. Concurrent yfinance downloads (snapshot price
# fetches racing fx-impact) have been observed returning another symbol's
# data — e.g. the USDCNY series filled with a JP stock price — which then
# poisons the 6h cache and blows up fx-impact by orders of magnitude.
# Reject the whole series if any close falls outside the band.
_FX_SANE = {"USD": (5.0, 10.0), "HKD": (0.6, 1.3), "JPY": (0.03, 0.09)}


def _fx_history(ccy: str, start: str) -> list[tuple[str, float]]:
    """Daily {ccy}CNY closes since start, sorted asc. Cached 6h; [] on failure."""
    from backend import persistent_cache as _pc
    key = f"fxhist:{ccy}:{start}"
    lo, hi = _FX_SANE.get(ccy, (0.0, float("inf")))
    cached = _pc.get(key)
    if cached is not None:
        if all(lo <= c <= hi for _, c in cached):
            return cached
        # poisoned pre-validation cache entry — drop and refetch
        _pc.delete(key)
    rows: list[tuple[str, float]] = []
    try:
        import yfinance as yf
        import pandas as pd
        end = (pd.Timestamp.now() + pd.DateOffset(days=1)).strftime("%Y-%m-%d")
        hist = yf.download(f"{ccy}CNY=X", start=start, end=end,
                           progress=False, auto_adjust=True)
        if hist is not None and not hist.empty:
            if isinstance(hist.columns, pd.MultiIndex):
                # select by explicit ticker label so a response carrying a
                # different symbol's columns raises instead of slipping through
                close = hist[("Close", f"{ccy}CNY=X")]
            else:
                close = hist["Close"]
            close = close.dropna()
            rows = [(pd.to_datetime(d).strftime("%Y-%m-%d"), float(c))
                    for d, c in zip(close.index, close)]
    except Exception:
        rows = []
    if rows and not all(lo <= c <= hi for _, c in rows):
        logger.warning("fx history %sCNY rejected: values out of sane range "
                       "[%s, %s] (yfinance cross-talk?)", ccy, lo, hi)
        rows = []
    if rows:
        _pc.put(key, rows, ttl=6 * 3600)
    return rows


@router.get("/fx-impact")
def get_fx_impact(user_id: str = Depends(get_current_user)):
    """YTD FX contribution to the CNY-denominated TWR.

    Per snapshot interval: fx return = sum of prev-day currency weights
    (equity MVs from market_data over net_assets) x FX moves; chained
    geometrically. Local (currency-hedged) return is the residual
    (1+total)/(1+fx)-1. Approximation: foreign-currency cash and margin
    loans are not in market_data, so their FX exposure is excluded.
    """
    _require_db()
    import datetime as _dt
    import bisect
    import json as _j
    from backend.services.portfolio_db import get_conn
    jan1 = f"{_dt.date.today().year}-01-01"
    with get_conn() as conn:
        snaps = conn.execute(
            "SELECT date, net_assets, capital, unit_nav, market_data, cash_json "
            "FROM daily_snapshots WHERE user_id=? AND date>=? "
            "AND net_assets IS NOT NULL ORDER BY date ASC",
            (user_id, jan1)).fetchall()
    series = []
    for date, na, cap, unav, md, cj in snaps:
        val = unav if (unav is not None and unav > 0) else (
            na / cap if cap and cap > 0 and na is not None else None)
        if val is not None and na and na > 0:
            series.append((date, val, na, md, cj))
    if len(series) < 2:
        return {}

    fx_start = (_dt.date.fromisoformat(series[0][0]) - _dt.timedelta(days=10)).isoformat()
    hist: dict[str, tuple[list[str], list[float]]] = {}
    for ccy in ("USD", "HKD", "JPY"):
        rows = _fx_history(ccy, fx_start)
        if rows:
            hist[ccy] = ([r[0] for r in rows], [r[1] for r in rows])
    if not hist:
        return {}

    def rate_on(ccy: str, date: str) -> float | None:
        if ccy == "CNY":
            return 1.0
        h = hist.get(ccy)
        if not h:
            return None
        i = bisect.bisect_right(h[0], date) - 1
        return h[1][i] if i >= 0 else None

    def ccy_weights(md_raw, cash_raw, date: str, net_assets: float) -> dict[str, float] | None:
        if not md_raw:
            return None
        try:
            md = _j.loads(md_raw)
        except Exception:
            return None
        out: dict[str, float] = {}
        for mkt, v in md.items():
            ccy = _MARKET_CCY.get(mkt, "USD")
            if isinstance(v, dict):
                # legacy nested {ccy: native MV} format (2026-03 early rows)
                for c, native in v.items():
                    r = rate_on(c, date)
                    if r is None or not isinstance(native, (int, float)):
                        return None
                    out[c] = out.get(c, 0.0) + native * r
            elif isinstance(v, (int, float)):
                out[ccy] = out.get(ccy, 0.0) + v
        # Signed per-currency cash/liability exposure (recorded since
        # 2026-07); folds margin loans in — exact from that date onward
        if cash_raw:
            try:
                for c, v in _j.loads(cash_raw).items():
                    if isinstance(v, (int, float)):
                        out[c] = out.get(c, 0.0) + v
            except Exception:
                pass
        return {c: mv / net_assets for c, mv in out.items()}

    tot_f = fx_f = 1.0
    covered = 0
    for i in range(1, len(series)):
        d0, v0, na0, md0, cj0 = series[i - 1]
        d1, v1 = series[i][0], series[i][1]
        tot_f *= v1 / v0
        w = ccy_weights(md0, cj0, d0, na0)
        if w is None:
            continue
        r_fx = 0.0
        ok = True
        for ccy, wt in w.items():
            if ccy == "CNY":
                continue
            r0, r1 = rate_on(ccy, d0), rate_on(ccy, d1)
            if not r0 or not r1:
                ok = False
                break
            r_fx += wt * (r1 / r0 - 1)
        if ok:
            fx_f *= 1 + r_fx
            covered += 1
    if covered == 0:
        return {}
    return {
        "total_pct": round((tot_f - 1) * 100, 2),
        "fx_pp": round((fx_f - 1) * 100, 2),
        "local_pct": round((tot_f / fx_f - 1) * 100, 2),
        "start": series[0][0],
        "end": series[-1][0],
    }


@router.get("/risk")
def get_risk(user_id: str = Depends(get_current_user)):
    """Leverage & margin stress view for a levered personal fund.

    Marks positions at cached live prices (cost fallback), takes signed
    cash by account/currency (negative = in-broker financing) plus
    margin_balances liabilities, and rakes the book over an equity-shock ×
    FX-shock grid. Coverage ratio = (broker MV + positive cash) / financing
    — the 维持担保比例 convention. Brokers' true maintenance requirements
    are position-specific, so thresholds are indicative, not exact.
    """
    path = _require_db()
    from backend.services.portfolio_db import get_conn
    from backend.services.portfolio_prices import (
        fetch_price, get_fx_rates as _get_fx, refresh_all_prices)
    fx = _get_fx(db_path=path)
    refresh_all_prices(db_path=path)
    with get_conn() as conn:
        pos_rows = conn.execute(
            "SELECT ticker, broker, currency, quantity, cost_price FROM positions "
            "WHERE status='open' AND user_id=? AND quantity > 0", (user_id,)).fetchall()
        cash_rows = conn.execute(
            "SELECT account, currency, balance FROM cash_balances WHERE user_id=?",
            (user_id,)).fetchall()
        margin_rows = conn.execute(
            "SELECT category, currency, amount FROM margin_balances WHERE user_id=?",
            (user_id,)).fetchall()

    positions = []
    for r in pos_rows:
        price, _ = fetch_price(r["ticker"])
        if price is None:
            price = r["cost_price"]
        positions.append({
            "broker": r["broker"], "ccy": r["currency"],
            "mv": r["quantity"] * price * fx.get(r["currency"], 1.0),
        })
    cash = [{"account": r["account"], "ccy": r["currency"],
             "cny": r["balance"] * fx.get(r["currency"], 1.0)} for r in cash_rows]
    # margin_balances rows are liabilities (positive amounts owed)
    liabs = [{"ccy": r["currency"], "category": r["category"],
              "cny": r["amount"] * fx.get(r["currency"], 1.0)} for r in margin_rows]

    def _shock(e: float, f: float) -> dict:
        """Apply equity shock e to all positions and FX shock f to every
        non-CNY exposure (assets and liabilities alike — a JPY loan hedges
        JPY stocks through the same factor)."""
        def fxm(ccy): return 1.0 + (f if ccy != "CNY" else 0.0)
        mv = {}
        for p in positions:
            v = p["mv"] * (1 + e) * fxm(p["ccy"])
            mv[p["broker"]] = mv.get(p["broker"], 0.0) + v
        pos_cash, fin = {}, {}
        for c in cash:
            v = c["cny"] * fxm(c["ccy"])
            if v >= 0:
                pos_cash[c["account"]] = pos_cash.get(c["account"], 0.0) + v
            else:
                fin[c["account"]] = fin.get(c["account"], 0.0) + (-v)
        liab_total = sum(l["cny"] * fxm(l["ccy"]) for l in liabs)
        total_mv = sum(mv.values())
        total_pos_cash = sum(pos_cash.values())
        total_fin = sum(fin.values())
        nav = total_mv + total_pos_cash - total_fin - liab_total
        debt = total_fin + liab_total
        brokers = []
        for b in sorted(set(list(mv) + list(fin) + list(pos_cash))):
            f_amt = fin.get(b, 0.0)
            collateral = mv.get(b, 0.0) + pos_cash.get(b, 0.0)
            brokers.append({
                "broker": b, "mv": mv.get(b, 0.0), "pos_cash": pos_cash.get(b, 0.0),
                "financing": f_amt,
                "coverage": (collateral / f_amt) if f_amt > 1 else None,
            })
        covs = [b["coverage"] for b in brokers if b["coverage"] is not None]
        return {"nav": nav, "debt": debt, "total_assets": total_mv + total_pos_cash,
                "debt_to_nav": debt / nav if nav > 0 else None,
                "gross_to_nav": (total_mv + total_pos_cash) / nav if nav > 0 else None,
                "worst_coverage": min(covs) if covs else None,
                "brokers": brokers}

    base = _shock(0.0, 0.0)

    # Per-broker distance to coverage thresholds (equity move only, FX flat):
    # (MV·(1+x) + posCash) / fin = T  →  x = (T·fin − posCash)/MV − 1
    for b in base["brokers"]:
        if b["coverage"] is not None and b["mv"] > 0:
            b["drop_to_150"] = min(0.0, (1.5 * b["financing"] - b["pos_cash"]) / b["mv"] - 1)
            b["drop_to_130"] = min(0.0, (1.3 * b["financing"] - b["pos_cash"]) / b["mv"] - 1)
        else:
            b["drop_to_150"] = b["drop_to_130"] = None

    # Net exposure by currency (assets − liabilities, signed cash folded in)
    ccy_net: dict[str, float] = {}
    for p in positions:
        ccy_net[p["ccy"]] = ccy_net.get(p["ccy"], 0.0) + p["mv"]
    for c in cash:
        ccy_net[c["ccy"]] = ccy_net.get(c["ccy"], 0.0) + c["cny"]
    for l in liabs:
        ccy_net[l["ccy"]] = ccy_net.get(l["ccy"], 0.0) - l["cny"]
    nav0 = base["nav"]
    exposure = [{"ccy": k, "net_cny": round(v), "pct_nav": round(v / nav0 * 100, 1) if nav0 > 0 else None}
                for k, v in sorted(ccy_net.items(), key=lambda kv: -abs(kv[1]))]

    grid = []
    for e in (0.0, -0.10, -0.20, -0.30, -0.40):
        for f in (-0.10, -0.05, 0.0, 0.05):
            s = _shock(e, f)
            grid.append({
                "equity_shock": e, "fx_shock": f,
                "nav": round(s["nav"]),
                "nav_pct": round((s["nav"] / nav0 - 1) * 100, 1) if nav0 > 0 else None,
                "debt_to_nav": round(s["debt_to_nav"], 3) if s["debt_to_nav"] is not None else None,
                "worst_coverage": round(s["worst_coverage"], 3) if s["worst_coverage"] is not None else None,
                # per-account coverage so the grid isn't a black box that
                # silently tracks whichever broker happens to be worst
                "coverages": {b["broker"]: round(b["coverage"], 3)
                              for b in s["brokers"] if b["coverage"] is not None},
            })

    for b in base["brokers"]:
        for k in ("mv", "pos_cash", "financing"):
            b[k] = round(b[k])
        if b["coverage"] is not None:
            b["coverage"] = round(b["coverage"], 3)

    return {
        "nav": round(base["nav"]),
        "total_assets": round(base["total_assets"]),
        "debt": round(base["debt"]),
        "off_exchange": round(sum(l["cny"] for l in liabs if l["category"] == "off_exchange")),
        "debt_to_nav": round(base["debt_to_nav"], 3) if base["debt_to_nav"] is not None else None,
        "gross_to_nav": round(base["gross_to_nav"], 3) if base["gross_to_nav"] is not None else None,
        "worst_coverage": base["worst_coverage"] and round(base["worst_coverage"], 3),
        "brokers": base["brokers"],
        "currency_exposure": exposure,
        "grid": grid,
    }


@router.get("/ibkr-recon")
def ibkr_recon(user_id: str = Depends(get_current_user)):
    """IBKR Flex reconciliation (review-only). {} when not configured.

    Env-gated single-owner feature: the Flex token lives in backend/.env,
    so this only activates on deployments whose owner configured it.
    """
    from backend.services import ibkr_flex
    if not ibkr_flex.enabled():
        return {}
    _require_db()
    from backend.services.portfolio_db import get_conn
    try:
        with get_conn() as conn:
            res = ibkr_flex.reconcile(conn, user_id=user_id)
        # None with the feature enabled = gateway down and no cached
        # statement — surface that instead of silently showing nothing
        return res or {"unavailable": True}
    except Exception as e:
        logger.warning("ibkr recon failed: %s", e)
        return {"unavailable": True}


class IbkrReconApplyItem(BaseModel):
    kind: str
    ticker: str


class IbkrReconApplyRequest(BaseModel):
    items: list[IbkrReconApplyItem]


@router.post("/ibkr-recon/apply")
def ibkr_recon_apply(req: IbkrReconApplyRequest,
                     user_id: str = Depends(get_current_user)):
    """One-click apply for mechanical recon diffs (cash / cost).

    Amounts are re-derived server-side from the current Flex statement —
    the client only names which diffs to apply. qty/missing_* are refused
    (semantic entries via the Trade panel keep attribution intact).
    """
    from backend.services import ibkr_flex
    if not ibkr_flex.enabled():
        raise HTTPException(status_code=400, detail="IBKR Flex not configured")
    _require_db()
    from backend.services.portfolio_db import get_conn
    with get_conn() as conn:
        with conn:  # transaction — commit on success
            res = ibkr_flex.apply_diffs(
                conn, [i.model_dump() for i in req.items], user_id=user_id)
    if res["applied"]:
        logger.info("ibkr recon applied %d diff(s) for %s: %s",
                    len(res["applied"]), user_id,
                    [(d["kind"], d["ticker"]) for d in res["applied"]])
    return res


@router.post("/ibkr-recon/ignore")
def ibkr_recon_ignore(req: IbkrReconApplyRequest,
                      user_id: str = Depends(get_current_user)):
    """Whitelist known intentional diffs (e.g. transfer-in cost conventions).

    Pinned to both sides' current values — if tracker or IBKR moves, the
    diff reappears in the banner.
    """
    from backend.services import ibkr_flex
    if not ibkr_flex.enabled():
        raise HTTPException(status_code=400, detail="IBKR Flex not configured")
    _require_db()
    from backend.services.portfolio_db import get_conn
    with get_conn() as conn:
        with conn:
            res = ibkr_flex.ignore_diffs(
                conn, [i.model_dump() for i in req.items], user_id=user_id)
    return res


@router.get("/dividends")
def list_dividends(user_id: str = Depends(get_current_user)):
    """Dividend ledger (broker-reported cash transactions) + per-ticker net.

    net = dividends + payments-in-lieu − withholding tax, converted to CNY
    at current rates (amounts are small relative to NAV; historical-rate
    precision isn't worth the bookkeeping).
    """
    path = _require_db()
    from backend.services.portfolio_db import get_conn
    from backend.services.ibkr_flex import _ensure_dividend_table
    from backend.services.portfolio_prices import get_fx_rates as _get_fx
    fx = _get_fx(db_path=path)
    with get_conn() as conn:
        _ensure_dividend_table(conn)
        rows = [dict(r) for r in conn.execute(
            "SELECT ticker, type, currency, amount, date, description, account "
            "FROM dividend_log WHERE user_id=? ORDER BY date DESC, id DESC",
            (user_id,))]
    by_ticker: dict[str, dict] = {}
    for r in rows:
        tk = r["ticker"] or "(cash)"
        g = by_ticker.setdefault(tk, {"net_native": 0.0, "currency": r["currency"], "net_cny": 0.0})
        g["net_native"] += r["amount"]
        g["net_cny"] += r["amount"] * fx.get(r["currency"] or "CNY", 1.0)
    for g in by_ticker.values():
        g["net_native"] = round(g["net_native"], 2)
        g["net_cny"] = round(g["net_cny"], 2)
    return {"rows": rows, "by_ticker": by_ticker}


@router.delete("/ibkr-recon/ignore")
def ibkr_recon_clear_ignores(user_id: str = Depends(get_current_user)):
    """Clear the whitelist — every ignored diff shows again."""
    from backend.services import ibkr_flex
    _require_db()
    from backend.services.portfolio_db import get_conn
    with get_conn() as conn:
        with conn:
            n = ibkr_flex.clear_ignores(conn, user_id=user_id)
    return {"cleared": n}


@router.post("/snapshot")
def record_snapshot(user_id: str = Depends(get_current_user)):
    """Day-0 snapshot: run the user's snapshot immediately after onboarding
    so the NAV curve starts today instead of tomorrow morning. Idempotent —
    daily_snapshots is INSERT OR IGNORE on (date, user_id)."""
    _require_db()
    from backend.services.portfolio_snapshot import take_snapshot
    try:
        res = take_snapshot(user_id=user_id, force=True)
        return {"ok": res is not None, "snapshot": res}
    except Exception as e:
        logger.warning("Day-0 snapshot failed for %s: %s", user_id, e)
        raise HTTPException(status_code=502, detail="Snapshot failed, will retry at the daily schedule")


# ── Import / Export ─────────────────────────────────────

_TEMPLATES = {
    "positions": {
        "headers": ["ticker", "name", "market", "broker", "currency", "quantity", "cost_price"],
        "rows": [
            ["600809.SS", "山西汾酒", "A股", "中信", "CNY", "200", "162.85"],
            ["GOOGL", "谷歌", "美股", "富途", "USD", "15", "635.1"],
        ],
        "filename": "positions_template.csv",
    },
    "cash": {
        "headers": ["account", "currency", "balance"],
        "rows": [
            ["中信", "CNY", "8057"],
            ["招商银行", "CNY", "9569"],
        ],
        "filename": "cash_template.csv",
    },
    "portfolio": {
        "headers": ["ticker", "name", "market", "broker", "currency", "quantity", "cost_price", "cash_balance"],
        "rows": [
            ["600809.SS", "山西汾酒", "A股", "中信", "CNY", "200", "162.85", ""],
            ["GOOGL", "谷歌", "美股", "富途", "USD", "15", "635.1", ""],
            ["", "", "", "中信", "CNY", "", "", "8057"],
            ["", "", "", "富途", "HKD", "", "", "35000"],
        ],
        "filename": "portfolio_template.csv",
    },
}


@router.get("/import-template/{template_type}")
def download_import_template(template_type: str):
    """Download a CSV template with headers and example rows."""
    if template_type not in _TEMPLATES:
        raise HTTPException(status_code=400, detail=f"Unknown template type: {template_type}. Use: {list(_TEMPLATES)}")
    tpl = _TEMPLATES[template_type]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(tpl["headers"])
    for row in tpl["rows"]:
        writer.writerow(row)
    # Encode with BOM for Excel compatibility
    content = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{tpl["filename"]}"'},
    )


@router.post("/import")
async def import_csv(file: UploadFile = File(...), user_id: str = Depends(get_current_user)):
    """Import positions or cash from a CSV file upload."""
    _require_db()
    from backend.services.portfolio_db import (
        init_db, get_conn, upsert_position, upsert_cash, upsert_account_setting,
    )
    init_db()

    # Read and decode file content (handle BOM)
    raw = await file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("utf-8")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="Empty or invalid CSV file")

    # Normalize headers (strip whitespace)
    headers = [h.strip().lower() for h in reader.fieldnames]

    # Detect type: mixed (has both ticker and cash_balance), positions-only, or cash-only
    has_ticker_col = "ticker" in headers
    has_cash_balance_col = "cash_balance" in headers
    has_account_col = "account" in headers

    if has_ticker_col:
        csv_type = "mixed" if has_cash_balance_col else "positions"
    elif has_account_col:
        csv_type = "cash"
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot detect CSV type from headers: {reader.fieldnames}. "
                   "Must have 'ticker' column (positions/mixed) or 'account' column (cash).",
        )

    imported_positions = 0
    imported_cash = 0
    accounts_created: list[str] = []
    errors: list[str] = []

    with get_conn() as conn:
        # Collect existing brokers/accounts
        existing_brokers = {
            r[0] for r in conn.execute("SELECT broker FROM account_settings WHERE user_id=?", (user_id,)).fetchall()
        }
        existing_cash_accounts = {
            r[0] for r in conn.execute("SELECT DISTINCT account FROM cash_balances WHERE user_id=?", (user_id,)).fetchall()
        }

        for i, raw_row in enumerate(reader, start=2):
            row = {k.strip().lower(): (v.strip() if v else "") for k, v in raw_row.items()}
            try:
                ticker = row.get("ticker", "").strip()
                cash_balance_str = row.get("cash_balance", "").strip()

                if csv_type == "cash":
                    # Pure cash CSV (account, currency, balance)
                    account = row["account"]
                    currency = row.get("currency", "CNY")
                    balance = float(row.get("balance", 0))
                    if not account:
                        errors.append(f"Row {i}: account is required")
                        continue
                    upsert_cash(conn, account=account, currency=currency, balance=balance, user_id=user_id)
                    imported_cash += 1

                elif ticker:
                    # Position row (has ticker)
                    name = row.get("name", "")
                    market = row.get("market", "")
                    broker = row.get("broker", "")
                    currency = row.get("currency", "CNY")
                    quantity = float(row.get("quantity", 0))
                    cost_price = float(row.get("cost_price", 0))

                    if not broker:
                        errors.append(f"Row {i}: broker is required for position rows")
                        continue

                    # Auto-create account_settings for new brokers
                    if broker not in existing_brokers:
                        upsert_account_setting(conn, broker=broker, capital_mode="cost", user_id=user_id)
                        existing_brokers.add(broker)
                        accounts_created.append(broker)

                    # Auto-create cash row for new broker (balance 0)
                    if broker not in existing_cash_accounts:
                        upsert_cash(conn, account=broker, currency="CNY", balance=0, user_id=user_id)
                        existing_cash_accounts.add(broker)

                    upsert_position(
                        conn, ticker=ticker, name=name, market=market,
                        broker=broker, currency=currency,
                        quantity=quantity, cost_price=cost_price, user_id=user_id,
                    )
                    imported_positions += 1

                elif cash_balance_str and csv_type == "mixed":
                    # Cash row in mixed CSV (ticker is empty, cash_balance has value)
                    account = row.get("broker", "")
                    currency = row.get("currency", "CNY")
                    balance = float(cash_balance_str)
                    if not account:
                        errors.append(f"Row {i}: broker is required for cash rows")
                        continue
                    upsert_cash(conn, account=account, currency=currency, balance=balance, user_id=user_id)
                    existing_cash_accounts.add(account)
                    imported_cash += 1

                else:
                    # Skip empty rows silently
                    pass

            except (ValueError, KeyError) as exc:
                errors.append(f"Row {i}: {exc}")

        conn.commit()

    # Detect similar broker names that might be duplicates
    warnings: list[str] = []
    all_brokers = list(existing_brokers)
    for i, a in enumerate(all_brokers):
        for b in all_brokers[i + 1:]:
            if a == b:
                continue
            # Check if one contains the other (e.g. "中信" vs "中信证券")
            if a in b or b in a:
                warnings.append(f"账户名 \"{a}\" 和 \"{b}\" 相似，是否为同一账户？如有误请在「设置」中统一。"
                                f" / Similar accounts \"{a}\" and \"{b}\" — same broker? Fix in Settings if needed.")

    imported = imported_positions + imported_cash
    result_type = "positions" if imported_cash == 0 else ("cash" if imported_positions == 0 else "mixed")
    return {
        "ok": True,
        "type": result_type,
        "imported": imported,
        "imported_positions": imported_positions,
        "imported_cash": imported_cash,
        "accounts_created": accounts_created,
        "errors": errors,
        "warnings": warnings,
    }


@router.get("/export")
def export_all(user_id: str = Depends(get_current_user)):
    """Export all portfolio data as JSON backup."""
    path = _require_db()
    return {
        "positions": _query(path, "SELECT * FROM positions WHERE user_id=? ORDER BY market, broker, name", (user_id,)),
        "cash_balances": _query(path, "SELECT * FROM cash_balances WHERE user_id=? ORDER BY account", (user_id,)),
        "account_settings": _query(path, "SELECT * FROM account_settings WHERE user_id=? ORDER BY broker", (user_id,)),
        "closed_trades": _query(path, "SELECT * FROM closed_trades WHERE user_id=? ORDER BY market, name", (user_id,)),
    }


@router.get("/export-excel")
def export_excel(user_id: str = Depends(get_current_user)):
    """Export portfolio data as a formatted Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers

    path = _require_db()

    positions = _query(path, """
        SELECT ticker, name, market, broker, currency, quantity, cost_price, status
        FROM positions WHERE user_id=? ORDER BY market, broker, name
    """, (user_id,))
    cash = _query(path, """
        SELECT account, currency, balance
        FROM cash_balances WHERE user_id=? ORDER BY account
    """, (user_id,))
    closed = _query(path, """
        SELECT ticker, name, market, broker, currency, quantity, cost_price,
               close_price, realized_pnl, realized_pnl_cny, close_date, notes
        FROM closed_trades WHERE user_id=? ORDER BY close_date DESC
    """, (user_id,))
    acct = _query(path, """
        SELECT broker, capital_mode, deposit_cny, deposit_fx, notes
        FROM account_settings WHERE user_id=? ORDER BY broker
    """, (user_id,))

    wb = Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(wrap_text=True, vertical="top")

    def _write_sheet(ws, title, headers, rows):
        ws.title = title
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_text
            cell.fill = header_fill
        for row in rows:
            ws.append([row.get(h) for h in headers])
        # Auto-fit columns
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 30)
        # Format numbers
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'

    # Sheet 1: Holdings
    h_headers = ["ticker", "name", "market", "broker", "currency", "quantity", "cost_price", "status"]
    _write_sheet(wb.active, "Holdings", h_headers, positions)

    # Sheet 2: Cash
    ws_cash = wb.create_sheet()
    _write_sheet(ws_cash, "Cash", ["account", "currency", "balance"], cash)

    # Sheet 3: Closed Trades
    ws_closed = wb.create_sheet()
    ct_headers = ["ticker", "name", "market", "broker", "currency", "quantity",
                  "cost_price", "close_price", "realized_pnl", "realized_pnl_cny", "close_date", "notes"]
    _write_sheet(ws_closed, "Closed Trades", ct_headers, closed)

    # Sheet 4: Account Settings
    ws_acct = wb.create_sheet()
    _write_sheet(ws_acct, "Accounts", ["broker", "capital_mode", "deposit_cny", "deposit_fx", "notes"], acct)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = __import__("datetime").date.today().isoformat()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Portfolio_{today}.xlsx"'},
    )


# ── Portfolio Event Feeds ──────────────────────────────────

import time
from urllib.request import urlopen, Request
import json as _json
from concurrent.futures import ThreadPoolExecutor, as_completed

_event_cache: dict[str, tuple[float, list]] = {}  # key → (timestamp, data)
_EVENT_TTL_NEWS = 600   # 10 min
_EVENT_TTL_OTHER = 3600  # 1 hour


def _cached_or_fetch(cache_key: str, ttl: int, fetch_fn):
    """Return cached data or call fetch_fn and cache result."""
    entry = _event_cache.get(cache_key)
    if entry and (time.time() - entry[0]) < ttl:
        return entry[1]
    data = fetch_fn()
    _event_cache[cache_key] = (time.time(), data)
    return data


def _fmp_get(path: str, apikey: str, params: str = "") -> list:
    """Quick FMP API GET, returns list or []."""
    if not apikey:
        return []
    url = f"https://financialmodelingprep.com/api/v3/{path}?apikey={apikey}{params}"
    try:
        req = Request(url, headers={"User-Agent": "ValueScope/1.0"})
        with urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read())
            return data if isinstance(data, list) else []
    except Exception as e:
        logger.warning(f"FMP API error ({path}): {e}")
        return []


def _get_tickers_by_market(db_path: str, user_id: str = "local") -> dict[str, list[dict]]:
    """Group open tickers by market type."""
    from backend.services.portfolio_db import get_conn, get_open_tickers
    conn = get_conn(db_path)
    tickers = get_open_tickers(conn, user_id=user_id)
    conn.close()
    groups: dict[str, list[dict]] = {"us": [], "hk": [], "ashare": [], "jp": []}
    for t in tickers:
        mk = (t.get("market") or "").upper()
        tk = t["ticker"]
        if mk in ("SHZ", "SHA", "BJ") or tk.endswith((".SZ", ".SS", ".BJ")):
            groups["ashare"].append(t)
        elif mk == "HKSE" or tk.endswith(".HK"):
            groups["hk"].append(t)
        elif tk.endswith(".T"):
            groups["jp"].append(t)
        else:
            groups["us"].append(t)
    return groups


@router.get("/news")
def portfolio_news(apikey: str = Query("", description="FMP API key"), user_id: str = Depends(get_current_user)):
    """Aggregated news for all portfolio holdings."""
    path = _require_db()
    apikey = apikey or os.environ.get("FMP_API_KEY", "")

    def _fetch():
        groups = _get_tickers_by_market(path, user_id=user_id)
        all_news = []

        # FMP news for US + HK stocks
        fmp_tickers = [t["ticker"] for t in groups["us"] + groups["hk"]]
        if fmp_tickers and apikey:
            # FMP accepts comma-separated symbols
            symbols = ",".join(fmp_tickers[:20])  # limit to 20
            items = _fmp_get("stock_news", apikey, f"&tickers={symbols}&limit=30")
            for item in items:
                all_news.append({
                    "title": item.get("title", ""),
                    "url": item.get("url", ""),
                    "source": item.get("site", ""),
                    "date": item.get("publishedDate", ""),
                    "ticker": item.get("symbol", ""),
                    "image": item.get("image", ""),
                })

        # A-share news via akshare
        for t in groups["ashare"][:10]:
            try:
                import akshare as ak
                code = t["ticker"].split(".")[0]
                df = ak.stock_news_em(symbol=code)
                if df is not None and not df.empty:
                    for _, row in df.head(3).iterrows():
                        all_news.append({
                            "title": str(row.get("新闻标题", "")),
                            "url": str(row.get("新闻链接", "")),
                            "source": str(row.get("文章来源", "东方财富")),
                            "date": str(row.get("发布时间", "")),
                            "ticker": t["ticker"],
                            "image": "",
                        })
            except Exception as e:
                logger.warning(f"akshare news error for {t['ticker']}: {e}")

        # Sort by date descending
        all_news.sort(key=lambda x: x.get("date", ""), reverse=True)
        return all_news[:50]

    return _cached_or_fetch(f"news:{path}:{user_id}", _EVENT_TTL_NEWS, _fetch)


@router.get("/earnings-calendar")
def portfolio_earnings(apikey: str = Query("", description="FMP API key"), user_id: str = Depends(get_current_user)):
    """Upcoming and recent earnings for portfolio holdings."""
    path = _require_db()
    apikey = apikey or os.environ.get("FMP_API_KEY", "")

    def _fetch():
        from datetime import datetime, timedelta
        groups = _get_tickers_by_market(path, user_id=user_id)
        events = []

        fmp_tickers = groups["us"] + groups["hk"] + groups["jp"]
        ticker_names = {t["ticker"]: t.get("name", t["ticker"]) for t in fmp_tickers}
        ticker_set = set(ticker_names.keys())

        if apikey and fmp_tickers:
            # Fetch calendar for a date range and filter for our tickers
            today = datetime.now()
            from_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
            to_date = (today + timedelta(days=90)).strftime("%Y-%m-%d")
            items = _fmp_get("earning_calendar", apikey,
                             f"&from={from_date}&to={to_date}")

            for item in items:
                sym = item.get("symbol", "")
                if sym not in ticker_set:
                    continue
                eps_est = item.get("epsEstimated")
                eps_act = item.get("eps")
                status = "reported" if eps_act is not None else "upcoming"
                events.append({
                    "ticker": sym,
                    "name": ticker_names.get(sym, sym),
                    "date": item.get("date", ""),
                    "eps_estimated": eps_est,
                    "eps_actual": eps_act,
                    "revenue_estimated": item.get("revenueEstimated"),
                    "revenue_actual": item.get("revenue"),
                    "status": status,
                })

        # Sort: upcoming first (by date asc), then reported (by date desc)
        upcoming = sorted([e for e in events if e["status"] == "upcoming"],
                          key=lambda x: x["date"])
        reported = sorted([e for e in events if e["status"] == "reported"],
                          key=lambda x: x["date"], reverse=True)
        return upcoming + reported

    return _cached_or_fetch(f"earnings:{path}:{user_id}", _EVENT_TTL_OTHER, _fetch)


@router.get("/rating-changes")
def portfolio_ratings(apikey: str = Query("", description="FMP API key"), user_id: str = Depends(get_current_user)):
    """Recent analyst rating changes for portfolio holdings."""
    path = _require_db()
    apikey = apikey or os.environ.get("FMP_API_KEY", "")

    def _fetch():
        groups = _get_tickers_by_market(path, user_id=user_id)
        changes = []

        fmp_tickers = groups["us"] + groups["hk"]
        if apikey and fmp_tickers:
            from datetime import datetime, timedelta
            cutoff = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")

            def _fetch_grades(t):
                items = _fmp_get(f"grade/{t['ticker']}", apikey, "&limit=10")
                result = []
                for item in items:
                    dt = item.get("date", "")[:10]
                    if dt < cutoff:
                        continue
                    prev = item.get("previousGrade", "")
                    new = item.get("newGrade", "")
                    direction = "maintain"
                    # Simple upgrade/downgrade detection
                    buy_words = {"buy", "overweight", "outperform", "strong buy", "positive"}
                    sell_words = {"sell", "underweight", "underperform", "strong sell", "negative"}
                    hold_words = {"hold", "neutral", "equal-weight", "market perform", "equal weight", "sector perform", "in-line"}
                    prev_l, new_l = prev.lower(), new.lower()
                    prev_rank = 3 if any(w in prev_l for w in buy_words) else (1 if any(w in prev_l for w in sell_words) else 2)
                    new_rank = 3 if any(w in new_l for w in buy_words) else (1 if any(w in new_l for w in sell_words) else 2)
                    if new_rank > prev_rank:
                        direction = "upgrade"
                    elif new_rank < prev_rank:
                        direction = "downgrade"

                    result.append({
                        "ticker": t["ticker"],
                        "name": t.get("name", t["ticker"]),
                        "date": dt,
                        "company": item.get("gradingCompany", ""),
                        "previous": prev,
                        "new": new,
                        "direction": direction,
                    })
                return result

            with ThreadPoolExecutor(max_workers=8) as pool:
                futures = {pool.submit(_fetch_grades, t): t for t in fmp_tickers[:20]}
                for f in as_completed(futures):
                    try:
                        changes.extend(f.result())
                    except Exception as e:
                        logger.warning(f"Rating fetch error: {e}")

        changes.sort(key=lambda x: x["date"], reverse=True)
        return changes

    return _cached_or_fetch(f"ratings:{path}:{user_id}", _EVENT_TTL_OTHER, _fetch)
