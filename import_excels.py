#!/usr/bin/env python3
# Copyright (c) 2025 Alan He. Licensed under MIT.
"""Import existing ValueScope Excel files into SQLite database.

Usage:
    python import_excels.py --db valuations.db
    python import_excels.py --db valuations.db --dir /path/to/excel/files
"""

import argparse
import os
import re
import sys

from openpyxl import load_workbook

# Add ValueScope to path so we can import db_export
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'valuescope'))
from modeling.db_export import save_to_db


def parse_filename(filename):
    """Extract metadata from Excel filename.

    Pattern: {company_name}_valuation_{YYYYMMDD}[_{engine}].xlsx
    """
    stem = os.path.splitext(filename)[0]
    match = re.match(r'^(.+?)_valuation_(\d{8})(?:_(.+))?$', stem)
    if not match:
        return None

    company_name = match.group(1)
    date_str = match.group(2)
    engine_raw = match.group(3)

    valuation_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
    ai_engine = engine_raw.replace('_', ' ') if engine_raw else None
    # Excel imports with AI engine → 'copilot' (can't distinguish auto vs copilot from filename)
    mode = 'copilot' if ai_engine else 'manual'

    return {
        'company_name': company_name,
        'valuation_date': valuation_date,
        'ai_engine': ai_engine,
        'mode': mode,
    }


def _safe_float(value, default=None):
    """Convert a cell value to float, returning default on failure."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def extract_from_excel(filepath):
    """Read valuation data from an Excel file."""
    wb = load_workbook(filepath, data_only=True)
    ws1 = wb['Input and sensitivity']
    ws6 = wb['Valuation output']

    # Parse reported_currency from ws6 header
    # e.g. "Tencent Holdings Limited - in CNY, millions"
    header = ws6.cell(1, 1).value or ''
    currency_match = re.search(r'in\s+(\w+),\s+millions', header)
    reported_currency = currency_match.group(1) if currency_match else ''

    # Read valuation parameters from Input sheet
    rg1 = _safe_float(ws1.cell(3, 2).value, 0)
    rg2 = _safe_float(ws1.cell(4, 2).value, 0)
    em = _safe_float(ws1.cell(6, 2).value, 0)
    tr = _safe_float(ws1.cell(14, 2).value, 0)
    wacc_val = _safe_float(ws1.cell(11, 2).value, 0)

    valuation_params = {
        'base_year': int(_safe_float(ws1.cell(2, 2).value, 0)),
        'ttm_quarter': '',
        'ttm_label': '',
        'revenue_growth_1': rg1 * 100 if abs(rg1) < 1 else rg1,
        'revenue_growth_2': rg2 * 100 if abs(rg2) < 1 else rg2,
        'ebit_margin': em * 100 if abs(em) < 1 else em,
        'convergence': _safe_float(ws1.cell(7, 2).value, 0),
        'revenue_invested_capital_ratio_1': _safe_float(ws1.cell(8, 2).value, 0),
        'revenue_invested_capital_ratio_2': _safe_float(ws1.cell(9, 2).value, 0),
        'revenue_invested_capital_ratio_3': _safe_float(ws1.cell(10, 2).value, 0),
        'tax_rate': tr * 100 if abs(tr) < 1 else tr,
        'wacc': wacc_val * 100 if abs(wacc_val) < 1 else wacc_val,
        'terminal_wacc': _safe_float(ws1.cell(12, 2).value, 0),
        'ronic': _safe_float(ws1.cell(13, 2).value, 0),
        'risk_free_rate': _safe_float(ws1.cell(5, 2).value, 0),
    }

    # Parse TTM label from ws6
    base_label = str(ws6.cell(1, 2).value or '')
    if 'TTM' in base_label:
        ttm_match = re.search(r'\((.+?)\)', base_label)
        if ttm_match:
            valuation_params['ttm_label'] = ttm_match.group(1)
            q_match = re.search(r'(Q\d)', ttm_match.group(1))
            if q_match:
                valuation_params['ttm_quarter'] = q_match.group(1)

    # DCF results -- extracted from formula cells (cached values)
    results = {
        'reported_currency': reported_currency,
        'pv_cf_next_10_years': _safe_float(ws6.cell(20, 2).value),
        'pv_terminal_value': _safe_float(ws6.cell(19, 2).value),
        'enterprise_value': _safe_float(ws6.cell(24, 2).value),
        'equity_value': _safe_float(ws6.cell(27, 2).value),
        'price_per_share': _safe_float(ws6.cell(29, 2).value),
        'cash': _safe_float(ws6.cell(22, 2).value),
        'total_investments': _safe_float(ws6.cell(23, 2).value),
        'total_debt': _safe_float(ws6.cell(25, 2).value),
        'minority_interest': _safe_float(ws6.cell(26, 2).value),
        'outstanding_shares': _safe_float(ws6.cell(28, 2).value),
    }

    company_profile = {
        'beta': _safe_float(ws1.cell(20, 2).value, 1.0),
        'country': None,
        'exchange': None,
        'currency': None,
        'marketCap': None,
        'price': None,
    }

    # Gap Analysis -- extract from sheet if it exists
    gap_analysis_result = None
    if 'Gap Analysis' in wb.sheetnames:
        ws_gap = wb['Gap Analysis']
        gap_analysis_result = _extract_gap_analysis(ws_gap)

    # AI raw text -- extract from sheet if it exists
    ai_result = None
    if 'Valuation Input Analysis' in wb.sheetnames:
        ws_ai = wb['Valuation Input Analysis']
        ai_text = _extract_sheet_text(ws_ai, start_row=3)
        if ai_text:
            ai_result = {'raw_text': ai_text, 'parameters': None}

    wb.close()
    return valuation_params, results, company_profile, gap_analysis_result, ai_result


def _extract_gap_analysis(ws):
    """Extract gap analysis data from the Gap Analysis sheet."""
    result = {}

    # Row 3: "当前股价: 522.50 HKD"
    price_text = str(ws.cell(3, 1).value or '')
    price_match = re.search(r'[\d,.]+', price_text.replace(',', ''))
    if price_match:
        result['current_price'] = float(price_match.group())

    # Row 4: "DCF 估值: 725.04 HKD"
    dcf_text = str(ws.cell(4, 1).value or '')
    dcf_match = re.search(r'[\d,.]+', dcf_text.replace(',', ''))
    if dcf_match:
        result['dcf_price'] = float(dcf_match.group())

    # Row 5: "差异: +38.8%"
    gap_text = str(ws.cell(5, 1).value or '')
    gap_match = re.search(r'[+-]?[\d.]+', gap_text)
    if gap_match:
        result['gap_pct'] = float(gap_match.group())

    # Row 6: "修正后估值: 650.00 HKD" (optional)
    adj_text = str(ws.cell(6, 1).value or '')
    adj_match = re.search(r'[\d,.]+', adj_text.replace(',', ''))
    if adj_match and '修正' in adj_text:
        result['adjusted_price'] = float(adj_match.group())

    # Full analysis text (rows 8+)
    result['analysis_text'] = _extract_sheet_text(ws, start_row=8)

    return result if result.get('current_price') else None


def _extract_sheet_text(ws, start_row=1):
    """Extract all text from a sheet starting at a given row."""
    lines = []
    for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            lines.append(str(val))
        else:
            lines.append('')
    # Trim trailing empty lines
    while lines and not lines[-1].strip():
        lines.pop()
    return '\n'.join(lines) if lines else None


def main():
    parser = argparse.ArgumentParser(description='Import ValueScope Excel files into SQLite')
    parser.add_argument('--db', required=True, help='Path to SQLite database file')
    parser.add_argument(
        '--dir',
        default=os.path.join(os.path.dirname(__file__), '..', 'valuescope', 'stock_valuation'),
        help='Directory containing Excel files (default: ../valuescope/stock_valuation)',
    )
    args = parser.parse_args()

    excel_dir = os.path.abspath(args.dir)
    if not os.path.isdir(excel_dir):
        print(f"Error: directory not found: {excel_dir}")
        sys.exit(1)

    files = [f for f in os.listdir(excel_dir) if f.endswith('.xlsx') and not f.startswith('~')]
    print(f"Found {len(files)} Excel files in {excel_dir}")

    imported = 0
    for f in sorted(files):
        meta = parse_filename(f)
        if not meta:
            print(f"  SKIP (cannot parse filename): {f}")
            continue

        filepath = os.path.join(excel_dir, f)
        try:
            vp, results, profile, gap, ai = extract_from_excel(filepath)
            save_to_db(
                db_path=os.path.abspath(args.db),
                ticker='',
                company_name=meta['company_name'],
                valuation_date=meta['valuation_date'],
                mode=meta['mode'],
                ai_engine=meta['ai_engine'],
                valuation_params=vp,
                results=results,
                company_profile=profile,
                gap_analysis_result=gap,
                ai_result=ai,
                source='excel_import',
            )
            imported += 1
            pps = results.get('price_per_share')
            pps_str = f", price/share={pps:,.2f}" if pps else ""
            print(f"  OK: {f}{pps_str}")
        except Exception as e:
            print(f"  ERROR: {f} -- {e}")

    print(f"\nDone. Imported {imported}/{len(files)} files into {os.path.abspath(args.db)}")


if __name__ == '__main__':
    main()
