# Copyright (c) 2025 Alan He. Licensed under MIT.
"""Excel export utilities for DCF valuation results."""

import io
import os
import re
import shutil
from .constants import TERMINAL_RISK_PREMIUM

# Paths are set once at import time by init_paths()
EXCEL_TEMPLATE_PATH = ''
EXCEL_OUTPUT_DIR = ''


def init_paths(project_root):
    """Set template and output directory paths based on project root.

    Output directory priority:
      1. VS_OUTPUT_DIR environment variable (user override)
      2. ../stock_valuation  (default — sibling directory next to project)
    """
    global EXCEL_TEMPLATE_PATH, EXCEL_OUTPUT_DIR
    EXCEL_TEMPLATE_PATH = os.path.join(project_root, 'modeling', 'DCF valuation template.xlsx')
    EXCEL_OUTPUT_DIR = os.environ.get(
        'VS_OUTPUT_DIR',
        os.path.join(project_root, '..', 'stock_valuation')
    )


def write_to_excel(filename, base_year_data, financial_data, valuation_params,
                   company_profile, total_equity_risk_premium,
                   gap_analysis_result=None, ai_result=None, wacc_sensitivity=None):
    """Write all valuation data to an Excel workbook.

    *filename* can be a file path (str) or an ``io.BytesIO`` object.
    When a BytesIO is passed the workbook is saved into it directly
    (no filesystem side-effects) so callers can use it for downloads.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    _to_stream = isinstance(filename, io.BytesIO)

    if _to_stream:
        wb = load_workbook(EXCEL_TEMPLATE_PATH)
    else:
        if not os.path.exists(EXCEL_OUTPUT_DIR):
            os.makedirs(EXCEL_OUTPUT_DIR)
        shutil.copy(EXCEL_TEMPLATE_PATH, filename)
        wb = load_workbook(filename)
    ws1 = wb['Input and sensitivity']
    ws2 = wb.create_sheet('Historical Financial Data')
    ws3 = wb.create_sheet('Income Statement')
    ws4 = wb.create_sheet('Balance Sheet')
    ws5 = wb.create_sheet('Cash Flow Statement')
    ws6 = wb['Valuation output']

    # ── Input parameters ──
    ws1.cell(row=2, column=2).value = valuation_params['base_year']
    ws1.cell(row=3, column=2).value = valuation_params['revenue_growth_1'] / 100
    ws1.cell(row=4, column=2).value = valuation_params['revenue_growth_2'] / 100
    ws1.cell(row=5, column=2).value = valuation_params['risk_free_rate']
    ws1.cell(row=6, column=2).value = valuation_params['ebit_margin'] / 100
    ws1.cell(row=7, column=2).value = valuation_params['convergence']
    ws1.cell(row=8, column=2).value = valuation_params['revenue_invested_capital_ratio_1']
    ws1.cell(row=9, column=2).value = valuation_params['revenue_invested_capital_ratio_2']
    ws1.cell(row=10, column=2).value = valuation_params['revenue_invested_capital_ratio_3']
    ws1.cell(row=11, column=2).value = valuation_params['wacc'] / 100
    ws1.cell(row=12, column=2).value = valuation_params['risk_free_rate'] + TERMINAL_RISK_PREMIUM
    ws1.cell(row=13, column=2).value = valuation_params['ronic']
    ws1.cell(row=14, column=2).value = valuation_params['tax_rate'] / 100

    # Update sensitivity table base values (E8 = revenue_growth_2, K2 = ebit_margin)
    ws1.cell(row=8, column=5).value = valuation_params['revenue_growth_2'] / 100   # E8
    ws1.cell(row=2, column=11).value = valuation_params['ebit_margin'] / 100        # K2

    ws1.cell(row=17, column=2).value = valuation_params['risk_free_rate']
    ws1.cell(row=18, column=2).value = base_year_data.get('Cost of Debt (%)', 0) / 100
    ws1.cell(row=19, column=2).value = total_equity_risk_premium
    ws1.cell(row=20, column=2).value = company_profile.get('beta', 1.0)

    # ── Historical data sheets ──
    for r in dataframe_to_rows(financial_data['summary'], index=True, header=True):
        ws2.append(r)

    # For A-shares, use complete raw financial statements; otherwise use FMP-compatible extracts
    income_export = financial_data.get('raw_income_statement', financial_data['income_statement'])
    balance_export = financial_data.get('raw_balance_sheet', financial_data['balance_sheet'])
    cashflow_export = financial_data.get('raw_cashflow_statement', financial_data['cashflow_statement'])

    for r in dataframe_to_rows(income_export, index=True, header=True):
        ws3.append(r)
    for r in dataframe_to_rows(balance_export, index=True, header=True):
        ws4.append(r)
    for r in dataframe_to_rows(cashflow_export, index=True, header=True):
        ws5.append(r)

    # ── Valuation output sheet ──
    company_name = company_profile.get('companyName', 'N/A')
    ws6.cell(row=1, column=1).value = f"{company_name} - in {base_year_data.get('Reported Currency', '')}, millions"

    ttm_label = valuation_params.get('ttm_label', '')
    if ttm_label:
        ws6.cell(row=1, column=2).value = f'Base ({ttm_label})'
    else:
        ws6.cell(row=1, column=2).value = 'Base Year'

    ws6.cell(row=3, column=2).value = float(base_year_data['Revenue Growth (%)']) / 100
    ws6.cell(row=4, column=2).value = float(base_year_data['Revenue'])
    ws6.cell(row=6, column=2).value = float(base_year_data['EBIT'])
    ws6.cell(row=9, column=2).value = float(base_year_data['Total Reinvestment'])
    ws6.cell(row=22, column=2).value = float(base_year_data['(-) Cash & Equivalents'])
    ws6.cell(row=23, column=2).value = float(base_year_data['(-) Total Investments'])
    ws6.cell(row=25, column=2).value = float(base_year_data['(+) Total Debt'])
    ws6.cell(row=26, column=2).value = float(base_year_data['Minority Interest'])
    ws6.cell(row=28, column=2).value = base_year_data['Outstanding Shares']
    ws6.cell(row=33, column=2).value = float(base_year_data['Invested Capital'])

    # ── Number format for summary data ──
    AMOUNT_ROWS = {'Revenue', 'EBIT',
                   '(+) Capital Expenditure', '(-) D&A', '(+) ΔWorking Capital', 'Total Reinvestment',
                   '(+) Total Debt', '(+) Total Equity', 'Minority Interest',
                   '(-) Cash & Equivalents', '(-) Total Investments', 'Invested Capital'}
    RATIO_ROWS = {'Revenue Growth (%)', 'EBIT Growth (%)', 'EBIT Margin (%)', 'Tax Rate (%)',
                  'Revenue / IC', 'Debt to Assets (%)', 'Cost of Debt (%)',
                  'ROIC (%)', 'ROE (%)', 'Dividend Yield (%)', 'Payout Ratio (%)'}
    for row in ws2.iter_rows(min_row=2):
        row_label = row[0].value
        if row_label in AMOUNT_ROWS:
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
        elif row_label in RATIO_ROWS:
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0'

    # ── Gap analysis sheet ──
    if gap_analysis_result and gap_analysis_result.get('analysis_text'):
        ws_gap = wb.create_sheet('Gap Analysis')
        currency = gap_analysis_result.get('currency', '')
        ws_gap.cell(row=1, column=1).value = 'DCF 估值 vs 当前股价 差异分析'
        ws_gap.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_gap.cell(row=3, column=1).value = f"当前股价: {gap_analysis_result['current_price']:.2f} {currency}"
        dcf_price_str = f"DCF 估值: {gap_analysis_result['dcf_price']:.2f} {currency}"
        if gap_analysis_result.get('dcf_price_raw') is not None:
            rcur = gap_analysis_result.get('reported_currency', '')
            frate = gap_analysis_result.get('forex_rate', 0)
            dcf_price_str += f"  ({gap_analysis_result['dcf_price_raw']:.2f} {rcur} × {frate:.4f})"
        ws_gap.cell(row=4, column=1).value = dcf_price_str
        ws_gap.cell(row=5, column=1).value = f"差异: {gap_analysis_result['gap_pct']:+.1f}%"
        if gap_analysis_result.get('adjusted_price') is not None:
            ws_gap.cell(row=6, column=1).value = f"修正后估值: {gap_analysis_result['adjusted_price']:,.2f} {currency}"

        analysis_text = re.sub(r'\n?\s*ADJUSTED_PRICE:.*$', '', gap_analysis_result['analysis_text']).strip()
        for i, line in enumerate(analysis_text.split('\n'), start=8):
            ws_gap.cell(row=i, column=1).value = line
        ws_gap.column_dimensions['A'].width = 120
        _wrap_align = Alignment(wrap_text=True, vertical='top')
        for row in ws_gap.iter_rows(min_row=1, max_col=1):
            for cell in row:
                cell.alignment = _wrap_align

    # ── WACC sensitivity on Input sheet ──
    if wacc_sensitivity:
        wacc_results, wacc_base = wacc_sensitivity
        wacc_start_row = 17
        ws1.cell(row=wacc_start_row, column=5).value = 'WACC Sensitivity Analysis'
        ws1.cell(row=wacc_start_row, column=5).font = Font(bold=True)
        ws1.cell(row=wacc_start_row + 1, column=5).value = 'WACC'
        ws1.cell(row=wacc_start_row + 2, column=5).value = 'Price / Share'
        for j, (wacc_val, price) in enumerate(wacc_results.items()):
            col_idx = 6 + j  # start from column F
            ws1.cell(row=wacc_start_row + 1, column=col_idx).value = wacc_val / 100
            ws1.cell(row=wacc_start_row + 1, column=col_idx).number_format = '0.0%'
            ws1.cell(row=wacc_start_row + 2, column=col_idx).value = price
            ws1.cell(row=wacc_start_row + 2, column=col_idx).number_format = '#,##0'
            if wacc_val == wacc_base:
                ws1.cell(row=wacc_start_row + 1, column=col_idx).font = Font(bold=True)
                ws1.cell(row=wacc_start_row + 2, column=col_idx).font = Font(bold=True)

    # ── AI analysis sheet ──
    if ai_result and ai_result.get('raw_text'):
        ws_ai = wb.create_sheet('Valuation Input Analysis')
        ws_ai.cell(row=1, column=1).value = 'AI 估值假设分析'
        ws_ai.cell(row=1, column=1).font = Font(bold=True, size=14)
        ai_text = ai_result['raw_text']
        for i, line in enumerate(ai_text.split('\n'), start=3):
            ws_ai.cell(row=i, column=1).value = line
        ws_ai.column_dimensions['A'].width = 120
        _wrap_align = Alignment(wrap_text=True, vertical='top')
        for row in ws_ai.iter_rows(min_row=1, max_col=1):
            for cell in row:
                cell.alignment = _wrap_align

    # ── Auto-fit column widths for data sheets ──
    for ws in [ws2, ws3, ws4, ws5]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)
